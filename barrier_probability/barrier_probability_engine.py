# -*- coding: utf-8 -*-

# PUBLIC-SAFE PORTFOLIO VERSION
# Real credentials, client data, private endpoints and internal paths were removed or replaced by placeholders.
"""
calcula_prob_barreira_v8.py

✅ Estrutura alinhada com sua planilha "Probabilidade de Rompimento.xlsx":
Abas usadas (e somente elas):
- Operações - DATA
- RTD - Export
- RTD - Opition Quotes
- Resultados
- Calendário B3

✅ Sem Excel COM: NÃO abre/não fecha Excel, então não trava nem "buga" planilha.
✅ O código apenas preenche os TICKERS nas abas RTD e depois (numa 2ª rodada) lê RTD - Export
   para calcular smile/skew simples e probabilidade de ROMPIMENTO POR FECHAMENTO (barreira DOWN).

Fluxo recomendado:
1) Rode o script (IDLE/F5). Ele preenche tickers nas abas RTD.
2) Abra o Excel, deixe o Profit preencher via RTD, aguarde IV/Delta/Strike/Vencimento aparecer, SALVE e feche.
3) Rode o script de novo (IDLE/F5). Ele calcula e grava a aba Resultados e (opcional) colunas na Operações - DATA.

Requisitos:
- openpyxl
- numpy
- requests (opcional, para puxar SELIC via BCB/SGS)

Observações:
- Para selecionar strikes sem depender do Spot (que pode estar vazio na 1ª rodada),
  o script usa um "spot proxy" = mediana dos strikes do vencimento escolhido, e aplica KO(%) sobre isso.
  Na 2ª rodada (com Spot real preenchido no Excel), o cálculo usa a barreira real.
"""

from __future__ import annotations

import os
import re
import math
import json
import sys
import argparse
import datetime as dt
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Set

import numpy as np
from openpyxl import load_workbook

try:
    import requests
except Exception:
    requests = None


# =========================
# Utils
# =========================

def _norm(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"\s+", " ", s)
    trans = str.maketrans({
        "ç":"c","ã":"a","á":"a","à":"a","â":"a",
        "é":"e","ê":"e","í":"i",
        "ó":"o","ô":"o","õ":"o",
        "ú":"u"
    })
    return s.translate(trans)

def _to_float(x):
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip()
    if s == "":
        return None
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None

def _to_date(x):
    if x is None:
        return None
    if isinstance(x, dt.datetime):
        return x.date()
    if isinstance(x, dt.date):
        return x
    s = str(x).strip()
    m = re.match(r"^(\d{2})/(\d{2})/(\d{4})$", s)
    if m:
        d, mo, y = map(int, m.groups())
        return dt.date(y, mo, d)
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", s)
    if m:
        y, mo, d = map(int, m.groups())
        return dt.date(y, mo, d)
    # yyyymmdd
    if re.match(r"^\d{8}$", s):
        try:
            return dt.datetime.strptime(s, "%Y%m%d").date()
        except Exception:
            return None
    return None

def _pct_to_frac(p):
    """Accepts 0.90, 90, '90,00%' -> 0.90"""
    if p is None:
        return None
    if isinstance(p, str) and "%" in p:
        p = p.replace("%", "")
    v = _to_float(p)
    if v is None:
        return None
    if v > 2.0:
        v = v / 100.0
    return v

def _first4_alnum(s: str) -> str:
    s = re.sub(r"[^A-Za-z0-9]", "", (s or ""))
    return s[:4].upper()

def _sgs_rate_aa(sgs_code: int = 11) -> float:
    """BCB/SGS última taxa (%a.a.). Série 11=SELIC diária (%a.a.). Retorna decimal."""
    if requests is None:
        return 0.0
    try:
        url = f"https://api.bcb.gov.br/dados/serie/bcdata.sgs.{sgs_code}/dados/ultimos/1?formato=json"
        r = requests.get(url, timeout=10)
        r.raise_for_status()
        j = r.json()
        if not j:
            return 0.0
        val = float(str(j[0]["valor"]).replace(",", "."))
        return val / 100.0
    except Exception:
        return 0.0


# =========================
# Prob (barreira por FECHAMENTO)
# =========================

BETA_DISCRETE = 0.5826  # Broadie-Glasserman-Kou continuity correction

def _N(z: float) -> float:
    return 0.5 * (1.0 + math.erf(z / math.sqrt(2.0)))

def prob_hit_down_continuous(S0: float, B: float, T: float, sigma: float, r: float=0.0, q: float=0.0) -> float:
    """P(min S_t <= B) under GBM, continuous monitoring."""
    if S0 <= 0 or B <= 0 or T <= 0 or sigma <= 0:
        return 0.0
    if B >= S0:
        return 1.0
    x0 = math.log(S0)
    a = math.log(B)
    mu = (r - q - 0.5 * sigma * sigma)
    sT = sigma * math.sqrt(T)
    z1 = (a - x0 - mu*T) / sT
    z2 = (a - x0 + mu*T) / sT
    expo = math.exp(2.0 * mu * (a - x0) / (sigma * sigma))
    p = _N(z1) + expo * _N(z2)
    return float(max(0.0, min(1.0, p)))

def prob_hit_down_close(S0: float, B: float, T: float, sigma: float, steps: int=252, r: float=0.0, q: float=0.0) -> float:
    """Approx. daily CLOSE monitoring via continuity correction."""
    if steps <= 0:
        steps = 252
    dt_ = T / steps
    if dt_ <= 0:
        return 0.0
    B_adj = B * math.exp(-BETA_DISCRETE * sigma * math.sqrt(dt_))
    return prob_hit_down_continuous(S0, B_adj, T, sigma, r=r, q=q)

def prob_below_at_T(S0: float, L: float, T: float, sigma: float, r: float=0.0, q: float=0.0) -> float:
    """P(S_T <= L)"""
    if S0 <= 0 or L <= 0 or T <= 0 or sigma <= 0:
        return 0.0
    mu = (r - q - 0.5 * sigma*sigma)
    z = (math.log(L/S0) - mu*T) / (sigma*math.sqrt(T))
    return float(_N(z))


# =========================
# SI_D_SEDE parsing (B3 séries)
# =========================

def iter_sede_lines(series_path: Path):
    if series_path.suffix.lower() == ".zip":
        import zipfile
        with zipfile.ZipFile(series_path) as z:
            names = [n for n in z.namelist() if n.lower().endswith(".txt")]
            if not names:
                names = z.namelist()
            raw = z.read(names[0])
        text = raw.decode("latin-1", errors="replace")
    else:
        text = series_path.read_text(encoding="latin-1", errors="replace")
    for line in text.splitlines():
        if line.startswith("02|"):
            yield line

def build_option_universe(series_path: Path) -> List[Dict]:
    """
    Retorna lista de opções com:
      root4, kind('P'/'C'), ticker, strike(float), expiry(int yyyymmdd)
    """
    out = []
    for line in iter_sede_lines(series_path):
        p = line.split("|")
        if len(p) < 18:
            continue
        kind_txt = (p[3] or "").strip().upper()   # "OPCAO DE COMPRA"/"OPCAO DE VENDA"
        root4 = _first4_alnum(p[6])              # underlying symbol
        ticker = (p[13] or "").strip().replace(" ", "")
        try:
            strike = float(p[16])
        except Exception:
            continue
        try:
            expiry = int(p[17])
        except Exception:
            continue
        kind = "P" if "VENDA" in kind_txt else ("C" if "COMPRA" in kind_txt else "")
        if not (root4 and ticker and kind):
            continue
        out.append({"root4": root4, "kind": kind, "ticker": ticker, "strike": strike, "expiry": expiry})
    return out


def pick_puts(universe: List[Dict], root4: str, target_date: dt.date,
              ko_frac: float, spot_hint: Optional[float],
              strike_window_pct: float=0.15, min_strikes: int=15) -> Tuple[Optional[dt.date], List[str], float]:
    """
    Escolhe:
      - expiry mais próximo de target_date (preferindo >=)
      - puts em strikes próximos de barreira

    Para funcionar mesmo com Spot vazio na 1ª rodada:
      - spot_proxy = mediana dos strikes do expiry escolhido
      - barreira_ref = (spot_hint if tiver) senão spot_proxy
      - target_strike = barreira_ref * ko_frac
    Retorna (expiry_date, tickers, spot_proxy)
    """
    root4 = (root4 or "").upper()
    puts = [u for u in universe if u["root4"] == root4 and u["kind"] == "P"]
    if not puts:
        return None, [], 0.0

    target = int(target_date.strftime("%Y%m%d"))
    expiries = sorted(set(u["expiry"] for u in puts))
    after = [e for e in expiries if e >= target]
    if after:
        expiry = min(after, key=lambda e: (e - target))
    else:
        expiry = min(expiries, key=lambda e: abs(e - target))

    cands = [u for u in puts if u["expiry"] == expiry]
    if not cands:
        return None, [], 0.0

    strikes = sorted(u["strike"] for u in cands)
    spot_proxy = float(np.median(strikes)) if strikes else 0.0
    ref = spot_hint if (spot_hint and spot_hint > 0) else spot_proxy
    if ref <= 0:
        ref = spot_proxy

    target_strike = ref * ko_frac
    low = target_strike * (1 - strike_window_pct)
    high = target_strike * (1 + strike_window_pct)

    near = [u for u in cands if low <= u["strike"] <= high]
    near.sort(key=lambda u: u["strike"])

    if len(near) < min_strikes:
        # widen window
        w = max(strike_window_pct, 0.25)
        low2 = target_strike * (1 - w)
        high2 = target_strike * (1 + w)
        near = [u for u in cands if low2 <= u["strike"] <= high2]
        near.sort(key=lambda u: u["strike"])

    if len(near) < min_strikes:
        # fallback: closest strikes
        cands.sort(key=lambda u: abs(u["strike"] - target_strike))
        near = sorted(cands[:min_strikes], key=lambda u: u["strike"])

    expiry_date = dt.datetime.strptime(str(expiry), "%Y%m%d").date()
    tickers = [u["ticker"] for u in near]
    return expiry_date, tickers, spot_proxy


# =========================
# Excel read/write
# =========================

def _get_headers(ws) -> Dict[str, int]:
    h = {}
    for c in range(1, ws.max_column+1):
        v = ws.cell(1, c).value
        if v is None:
            continue
        h[_norm(str(v))] = c
    return h

def _col(headers: Dict[str,int], *names: str) -> Optional[int]:
    for n in names:
        k = _norm(n)
        if k in headers:
            return headers[k]
    # fuzzy contains
    for k, idx in headers.items():
        for n in names:
            if _norm(n) in k:
                return idx
    return None

def read_operations(excel_path: Path) -> List[Dict]:
    wb = load_workbook(excel_path, data_only=True)
    ws = wb["Operações - DATA"]
    headers = _get_headers(ws)

    c_cod = _col(headers, "Código", "Codigo")
    c_ativo = _col(headers, "Ativo")
    c_fix = _col(headers, "Data de Fixing", "Fixing")
    c_venc = _col(headers, "Data de vencimento", "Vencimento")
    c_ko = _col(headers, "Knockout (%)", "Barreira (%)", "Knockout")
    c_spot = _col(headers, "Spot S0", "Spot")

    if not c_ativo or not c_fix or not c_ko:
        raise RuntimeError("Colunas mínimas não encontradas em Operações - DATA: Ativo, Data de Fixing, Knockout (%).")

    ops = []
    for r in range(2, ws.max_row+1):
        ativo = ws.cell(r, c_ativo).value
        if not ativo:
            continue
        ativo = str(ativo).strip().upper()
        root4 = _first4_alnum(ativo)
        fixing = _to_date(ws.cell(r, c_fix).value)
        venc = _to_date(ws.cell(r, c_venc).value) if c_venc else None
        ko_frac = _pct_to_frac(ws.cell(r, c_ko).value)
        spot = _to_float(ws.cell(r, c_spot).value) if c_spot else None
        codigo = ws.cell(r, c_cod).value if c_cod else None
        if not fixing or not ko_frac:
            continue
        ops.append({
            "row": r,
            "codigo": codigo,
            "ativo": ativo,
            "root4": root4,
            "fixing": fixing,
            "target_date": venc or fixing,
            "ko_frac": ko_frac,
            "spot": spot
        })
    return ops


def write_tickers_to_rtd_sheets(excel_path: Path, option_tickers: List[str]):
    """
    Preenche a coluna A das tabelas de:
      - RTD - Export   (coluna "Ativo")
      - RTD - Opition Quotes (coluna "Underlying" — na prática é ticker da opção)
    Remove qualquer repetição e limpa sobras.
    """
    wb = load_workbook(excel_path)
    tks = [t.strip().upper().replace(" ", "") for t in option_tickers if t and str(t).strip()]
    # distinct preserving order
    seen = set()
    distinct = []
    for t in tks:
        if t not in seen:
            distinct.append(t)
            seen.add(t)

    def fill_sheet(sheet_name: str, first_col_header: str):
        ws = wb[sheet_name]
        # find first column by header name (row 1)
        headers = _get_headers(ws)
        c0 = _col(headers, first_col_header)
        if not c0:
            raise RuntimeError(f"Não achei a coluna '{first_col_header}' na aba '{sheet_name}'.")
        # clear existing values below header
        for r in range(2, ws.max_row+1):
            ws.cell(r, c0).value = None
        # write tickers
        for i, t in enumerate(distinct, start=2):
            ws.cell(i, c0).value = t

    fill_sheet("RTD - Export", "Ativo")
    fill_sheet("RTD - Opition Quotes", "Underlying")
    wb.save(excel_path)


def read_rtd_export(excel_path: Path) -> List[Dict]:
    """
    Lê RTD - Export com headers como no seu print:
    Ativo | Vencimento | Strike | Of. Compra | Of. Venda | Volume | Cont. Abertos | VI Bid | VI Ask | Volt. Implícita - Opções | Delta
    """
    wb = load_workbook(excel_path, data_only=True)
    ws = wb["RTD - Export"]
    headers = _get_headers(ws)

    c_asset = _col(headers, "Ativo", "Asset")
    if not c_asset:
        return []

    c_exp = _col(headers, "Vencimento", "VEN")
    c_strike = _col(headers, "Strike", "PEX")
    c_bid = _col(headers, "Of. Compra", "OCP", "Bid")
    c_ask = _col(headers, "Of. Venda", "OVD", "Ask")
    c_vol = _col(headers, "Volume", "VOL")
    c_oi = _col(headers, "Cont. Abertos", "CAB", "OI")
    c_vib = _col(headers, "VI Bid", "VIB")
    c_via = _col(headers, "VI Ask", "VIA")
    c_iv = _col(headers, "Volt. Implícita - Opções", "Volt. Implícita", "IMPVT")
    c_delta = _col(headers, "Delta", "DELTA")

    out = []
    for r in range(2, ws.max_row+1):
        asset = ws.cell(r, c_asset).value
        if not asset:
            continue
        asset = str(asset).strip().upper().replace(" ", "")
        root4 = _first4_alnum(asset)

        exp = _to_date(ws.cell(r, c_exp).value) if c_exp else None
        strike = _to_float(ws.cell(r, c_strike).value) if c_strike else None

        bid = _to_float(ws.cell(r, c_bid).value) if c_bid else None
        ask = _to_float(ws.cell(r, c_ask).value) if c_ask else None
        vol = _to_float(ws.cell(r, c_vol).value) if c_vol else None
        oi = _to_float(ws.cell(r, c_oi).value) if c_oi else None

        vib = _to_float(ws.cell(r, c_vib).value) if c_vib else None
        via = _to_float(ws.cell(r, c_via).value) if c_via else None
        iv_any = _to_float(ws.cell(r, c_iv).value) if c_iv else None
        delta = _to_float(ws.cell(r, c_delta).value) if c_delta else None

        iv = None
        if vib is not None and via is not None:
            iv = 0.5*(vib+via)
        elif iv_any is not None:
            iv = iv_any

        if iv is not None and iv > 2.0:
            iv = iv/100.0

        out.append({
            "asset": asset,
            "root4": root4,
            "expiry": exp,
            "strike": strike,
            "bid": bid, "ask": ask,
            "vol": vol, "oi": oi,
            "iv": iv,
            "delta": delta
        })
    return out


def read_calendar_b3(excel_path: Path) -> Set[dt.date]:
    """
    Se Calendário B3 estiver preenchido com:
      Col A: Data
      Col B: Dia útil? (1/TRUE/Sim)
    retorna set de datas úteis.
    """
    wb = load_workbook(excel_path, data_only=True)
    ws = wb["Calendário B3"]
    headers = _get_headers(ws)
    c_date = _col(headers, "Data")
    c_biz = _col(headers, "Dia útil?", "Dia util?")
    if not c_date or not c_biz:
        return set()
    biz = set()
    for r in range(2, ws.max_row+1):
        d = _to_date(ws.cell(r, c_date).value)
        if not d:
            continue
        flag = ws.cell(r, c_biz).value
        ok = False
        if isinstance(flag, bool):
            ok = flag
        else:
            s = str(flag).strip().lower()
            ok = s in ("1","true","sim","s","yes","y")
        if ok:
            biz.add(d)
    return biz


# =========================
# Smile fit (sigma at barrier)
# =========================

def fit_smile_sigma_at_level(rows: List[Dict], root4: str, expiry: dt.date, level: float,
                             min_oi: float=50, min_vol: float=10, max_spread_pct: float=0.50,
                             delta_min: float=-0.60, delta_max: float=-0.05) -> Tuple[Optional[float], Dict]:
    use = [r for r in rows if r["root4"] == root4 and r["expiry"] == expiry and r["iv"] is not None and r["strike"]]
    if not use:
        return None, {"reason": "sem_quotes_para_expiry"}

    filt = []
    for r in use:
        bid, ask = r.get("bid"), r.get("ask")
        if bid and ask and bid > 0 and ask > 0:
            spread = (ask-bid) / max(1e-9, 0.5*(ask+bid))
            if spread > max_spread_pct:
                continue
        if r.get("oi") is not None and r["oi"] < min_oi:
            continue
        if r.get("vol") is not None and r["vol"] < min_vol:
            continue
        if r.get("delta") is not None:
            if not (delta_min <= r["delta"] <= delta_max):
                continue
        filt.append(r)

    if len(filt) < 5:
        filt = use

    strikes = np.array([r["strike"] for r in filt], dtype=float)
    ivs = np.array([r["iv"] for r in filt], dtype=float)
    x = np.log(strikes / max(1e-9, level))
    w = np.exp(- (x/0.15)**2 )
    try:
        coef = np.polyfit(x, ivs, deg=2, w=w)
        a, b, c = coef
        sigma = float(c)  # at x=0 => K=level
        sigma = max(0.01, min(5.0, sigma))
        return sigma, {"n": int(len(filt)), "coef": [float(a), float(b), float(c)]}
    except Exception as e:
        return None, {"reason": f"polyfit_falhou: {e}"}


# =========================
# Results write
# =========================

def write_results(excel_path: Path, ops: List[Dict], results: List[Dict]):
    wb = load_workbook(excel_path)
    ws = wb["Resultados"]
    # clear and rewrite
    for r in range(2, ws.max_row+1):
        for c in range(1, 50):
            ws.cell(r, c).value = None

    # header (row 1 assumed already ok)
    # find header mapping
    headers = _get_headers(ws)
    cols = {
        "Codigo": _col(headers, "Codigo", "Código"),
        "Ativo": _col(headers, "Ativo"),
        "Spot": _col(headers, "Spot"),
        "KO(%)": _col(headers, "KO(%)"),
        "Barreira(R$)": _col(headers, "Barreira(R$)"),
        "Fixing": _col(headers, "Fixing"),
        "Expiry": _col(headers, "Expiry"),
        "sigma": _col(headers, "sigma"),
        "P_hit_close": _col(headers, "P_hit_close"),
        "P_term": _col(headers, "P_term"),
        "r": _col(headers, "r"),
        "smile_info": _col(headers, "smile_info"),
    }
    # fallback positions if missing
    def col_or(i, default):
        return i if i else default

    # write rows
    out_r = 2
    for res in results:
        ws.cell(out_r, col_or(cols["Codigo"], 1)).value = res.get("codigo")
        ws.cell(out_r, col_or(cols["Ativo"], 2)).value = res.get("ativo")
        ws.cell(out_r, col_or(cols["Spot"], 3)).value = res.get("spot")
        ws.cell(out_r, col_or(cols["KO(%)"], 4)).value = res.get("ko")
        ws.cell(out_r, col_or(cols["Barreira(R$)"], 5)).value = res.get("barreira")
        ws.cell(out_r, col_or(cols["Fixing"], 6)).value = res.get("fixing")
        ws.cell(out_r, col_or(cols["Expiry"], 7)).value = res.get("expiry")
        ws.cell(out_r, col_or(cols["sigma"], 8)).value = res.get("sigma")
        ws.cell(out_r, col_or(cols["P_hit_close"], 9)).value = res.get("p_hit")
        ws.cell(out_r, col_or(cols["P_term"], 10)).value = res.get("p_term")
        ws.cell(out_r, col_or(cols["r"], 11)).value = res.get("r")
        ws.cell(out_r, col_or(cols["smile_info"], 12)).value = json.dumps(res.get("smile_info", {}), ensure_ascii=False)
        out_r += 1

    wb.save(excel_path)


# =========================
# Main
# =========================

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--series_file", required=True)
    parser.add_argument("--strike_window_pct", type=float, default=0.15)
    parser.add_argument("--min_strikes", type=int, default=15)
    parser.add_argument("--min_oi", type=float, default=50)
    parser.add_argument("--min_vol", type=float, default=10)
    parser.add_argument("--max_spread_pct", type=float, default=0.50)
    parser.add_argument("--delta_min", type=float, default=-0.60)
    parser.add_argument("--delta_max", type=float, default=-0.05)
    parser.add_argument("--sgs_code", type=int, default=11)
    args = parser.parse_args()

    excel_path = Path(args.workbook)
    series_path = Path(args.series_file)

    if not excel_path.exists():
        raise SystemExit(f"Workbook não encontrado: {excel_path}")
    if not series_path.exists():
        raise SystemExit(f"Series file não encontrado: {series_path}")

    # Load operations + universe
    ops = read_operations(excel_path)
    if not ops:
        print("❌ Nenhuma operação válida encontrada (verifique Ativo, Fixing e Knockout(%)).")
        return

    universe = build_option_universe(series_path)

    # 1) Build option tickers list from operations (even if Spot is not ready)
    all_option_tickers: List[str] = []
    chosen: Dict[int, Dict] = {}

    for op in ops:
        expiry_date, tickers, spot_proxy = pick_puts(
            universe, op["root4"], op["target_date"], op["ko_frac"], op.get("spot"),
            strike_window_pct=args.strike_window_pct, min_strikes=args.min_strikes
        )
        if not tickers:
            continue
        chosen[op["row"]] = {"expiry": expiry_date, "spot_proxy": spot_proxy}
        all_option_tickers.extend(tickers)

    if not all_option_tickers:
        print("❌ Não consegui selecionar opções a partir do SI_D_SEDE para os ativos das operações.")
        print("   Verifique se o SI_D_SEDE é o arquivo completo e se os códigos de ativo (4 letras) batem (ex.: PETR, VALE, etc.).")
        return

    # Always write tickers to RTD sheets (1ª rodada)
    write_tickers_to_rtd_sheets(excel_path, all_option_tickers)
    print(f"✅ Preenchi {len(set(all_option_tickers))} tickers de OPÇÕES em:")
    print("   - RTD - Export (coluna Ativo)")
    print("   - RTD - Opition Quotes (coluna Underlying)")

    # 2) Try to compute results if RTD - Export already has data
    quotes = read_rtd_export(excel_path)

    # detect if RTD populated: need at least some rows with expiry, strike and iv
    populated = any(r.get("expiry") and r.get("strike") and r.get("iv") for r in quotes)
    if not populated:
        print("⚠️  RTD - Export ainda não está preenchida (ou não foi salva com valores).")
        print("➡️  Abra o Excel, aguarde o RTD do Profit preencher (Strike/Vencimento/IV/Delta/OI/Volume), SALVE e feche.")
        print("➡️  Depois rode o script novamente (IDLE/F5) para calcular as probabilidades.")
        return

    # calendar
    biz = read_calendar_b3(excel_path)

    r_aa = _sgs_rate_aa(args.sgs_code)

    results_out: List[Dict] = []
    today = dt.date.today()

    # 3) Compute per operation
    for op in ops:
        spot = op.get("spot")
        # If Spot still missing, use proxy from chosen
        proxy = chosen.get(op["row"], {}).get("spot_proxy", 0.0)
        S0 = spot if (spot and spot > 0) else proxy
        if not S0 or S0 <= 0:
            continue

        barrier = S0 * op["ko_frac"]

        expiry_date = chosen.get(op["row"], {}).get("expiry")
        if not expiry_date:
            # fallback to nearest expiry in quotes
            exp_set = sorted({q["expiry"] for q in quotes if q["root4"] == op["root4"] and q["expiry"]})
            if exp_set:
                expiry_date = min(exp_set, key=lambda d: abs((d - op["target_date"]).days))
        if not expiry_date:
            continue

        sigma, sinfo = fit_smile_sigma_at_level(
            quotes, op["root4"], expiry_date, barrier,
            min_oi=args.min_oi, min_vol=args.min_vol, max_spread_pct=args.max_spread_pct,
            delta_min=args.delta_min, delta_max=args.delta_max
        )
        if sigma is None:
            continue

        # T in years: prefer business days if calendar provided, else calendar days
        if biz:
            # count business dates in (today, fixing] that are marked as business
            d0 = min(today, op["fixing"])
            d1 = max(today, op["fixing"])
            days = sum(1 for d in biz if d0 < d <= d1)
            days = max(1, days)
        else:
            days = max(1, (op["fixing"] - today).days)

        T = days / 252.0

        p_hit = prob_hit_down_close(S0, barrier, T, sigma, steps=252, r=r_aa, q=0.0)
        p_term = prob_below_at_T(S0, barrier, T, sigma, r=r_aa, q=0.0)

        results_out.append({
            "codigo": op.get("codigo"),
            "ativo": op.get("ativo"),
            "spot": float(S0),
            "ko": float(op["ko_frac"]),
            "barreira": float(barrier),
            "fixing": op["fixing"].isoformat(),
            "expiry": expiry_date.isoformat(),
            "sigma": float(sigma),
            "p_hit": float(p_hit),
            "p_term": float(p_term),
            "r": float(r_aa),
            "smile_info": sinfo
        })

    write_results(excel_path, ops, results_out)
    print(f"✅ Probabilidades calculadas: {len(results_out)} operações (aba Resultados).")
    print("   Modelo: barreira DOWN por FECHAMENTO (monitoramento diário).")


if __name__ == "__main__":
    # IDLE-friendly auto-detect
    if len(sys.argv) == 1:
        here = Path(__file__).resolve().parent
        # workbook: prefer the file name user uploaded
        wb = None
        for name in ["Probabilidade de Rompimento.xlsx", "Modelo_Prob_Barreira_clean.xlsx"]:
            p = here / name
            if p.exists():
                wb = p
                break
        if wb is None:
            cand = list(here.glob("*.xlsx"))
            wb = cand[0] if cand else None

        sf = None
        cand = list(here.glob("SI_D_SEDE*.zip")) + list(here.glob("SI_D_SEDE*.txt"))
        sf = cand[0] if cand else None

        if wb is None or sf is None:
            print("❌ Não encontrei automaticamente o workbook e/ou SI_D_SEDE na mesma pasta do .py.")
            print('Use via terminal: py calcula_prob_barreira_v8.py --workbook "C:\\path\\to\\Probabilidade_de_Rompimento.xlsx" --series_file "C:\\path\\to\\SI_D_SEDE.zip"')
            sys.exit(1)

        sys.argv += ["--workbook", str(wb), "--series_file", str(sf)]
        print(f"Modo IDLE: workbook = {wb}")
        print(f"Modo IDLE: series_file = {sf}")

    main()
