# -*- coding: utf-8 -*-

# PUBLIC-SAFE PORTFOLIO VERSION
# Real credentials, client data, private endpoints and internal paths were removed or replaced by placeholders.
"""
Early Unwind Automation – v2.5 (Public-Safe)

Principais ajustes desta versão:
- Mantém lógica de Soma_Bid_Offer IGUAL ao modelo antigo:
  soma de "Bid(+)/Offer(-)" por OPERAÇÃO (grupo) e só depois reduz para 1 linha.
- Chave de operação (op_key) prioriza (Conta + Ativo + Estrutura + Data_Operação + Fixing + Quantidade + Ref_Entrada quando existir).
  Se faltar algum campo, usa o que tiver; em último caso cai em "Código do Produto".
- Calcula CDI do período (Data_Operação -> hoje) via API do BCB/SGS (série 12, CDI diário em % ao dia)
  e inclui no corpo do e-mail as operações onde Resultado % > %CDI do período.
- Inclui coluna "Assessor" em todas as abas (a partir da "Advisor Base", se encontrada).
- Na aba "Antecipacao com Ganho" remove as colunas:
  Codigo_Produto, Barreira Knock In, Barreira Knock Out, KnockInAtingido, Preço de Ex Knock In
- Leitura resiliente: se Excel travar o arquivo (PermissionError), tenta salvar uma cópia via COM (Excel) e ler a cópia.

Requisitos:
- Python 3.10+
- pandas, openpyxl
- (opcional) pywin32 para email Outlook e para contornar arquivo travado em Excel
"""

# =========================
# BOOTSTRAP OneDrive/Excel
# - força diretório de trabalho na pasta do script
# - cria pasta de logs e grava traceback em caso de erro
# =========================
from pathlib import Path as _Path
import os as _os
import sys as _sys
import traceback as _traceback
from datetime import datetime as _dt_datetime

def _get_base_dir() -> _Path:
    if getattr(_sys, "frozen", False):
        return _Path(_sys.executable).resolve().parent
    return _Path(__file__).resolve().parent

BASE_DIR = _get_base_dir()
try:
    _os.chdir(BASE_DIR)
except Exception:
    pass

def _infer_root_dir(base_dir: _Path) -> _Path:
    name = (base_dir.name or "").strip().lower()
    if name in {"scripts py", "scripts_py", "scripts python", "scripts_python", "scripts"}:
        return base_dir.parent
    return base_dir

ROOT_DIR = _infer_root_dir(BASE_DIR)

# Pastas padrão (relativas ao ROOT_DIR):
#   ROOT_DIR\Scripts py        -> onde ficam os .py
#   ROOT_DIR\logs              -> logs centralizados
#   ROOT_DIR\Planilhas base    -> onde ficam as planilhas base baixadas
PLANILHAS_DIR = ROOT_DIR / "Planilhas base"

LOG_DIR = ROOT_DIR / "logs"
if not LOG_DIR.exists():
    # fallback para manter compatibilidade com versões antigas
    LOG_DIR = BASE_DIR / "_logs"
try:
    LOG_DIR.mkdir(parents=True, exist_ok=True)
except Exception:
    pass
def _log_unhandled_exception(prefix: str = "run_err") -> None:
    try:
        ts = _dt_datetime.now().strftime("%Y%m%d_%H%M%S")
        p_latest = LOG_DIR / f"{prefix}_LATEST.log"
        p_ts = LOG_DIR / f"{prefix}_{ts}.log"
        tb = _traceback.format_exc()
        p_latest.write_text(tb, encoding="utf-8", errors="ignore")
        p_ts.write_text(tb, encoding="utf-8", errors="ignore")
    except Exception:
        # não deixa o logging derrubar a execução
        pass


import os
import re
import sys
import time
import json
import shutil
import tempfile
import datetime as _dt
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

# =========================
# Configurações
# =========================

NOME_PASTA_SAIDA = "ANTECIPACAO"
NOME_ARQUIVO_SAIDA = "Resultados_Anticipacao.xlsx"

# Caso queira forçar nomes/caminhos (sem buscar por padrão), preencha:
FORCAR_ARQUIVO_MONITOR: Optional[str] = None
FORCAR_ARQUIVO_CUSTODIA: Optional[str] = None
FORCAR_ARQUIVO_ADVISOR_BASE: Optional[str] = None

# Padrões aceitos para descobrir os arquivos (case-insensitive, ignora acentos e símbolos).
PADROES_MONITOR = [
    "monitor saidas", "monitor_saidas", "monitor saida", "monitor_saidas", "monitor de saidas"
]
PADROES_CUSTODIA = [
    "acompanhamento de custodia", "acompanhamento-de-custodia", "acompanhamento custodia",
    "custodia", "acompanhamento de custódia"
]
PADROES_ADVISOR_BASE = [
    "advisor base", "advisor_base", "base assessores", "base assessoria", "base clientes"
]

# Email (opcional)
ENVIAR_EMAIL = True
EMAIL_TO = os.getenv("OPERATIONS_EMAIL", "operations@example.com")
EMAIL_SUBJECT = "Relatório de Operações - Antecipação"

# Assinatura (opcional) – procure no Desktop por "Assinatura.png" ou defina caminho completo aqui
ASSINATURA_FIXA: Optional[str] = None  # ex.: r"C:\Users\Public\Desktop\signature.png"

# CDI (BCB/SGS) – Série 12 (CDI % ao dia)
BCB_SGS_SERIE_CDI_DIA = 12
BCB_TIMEOUT_SEC = 15

# Títulos de saída (sem underscore nos cabeçalhos do Excel/e-mail)
OUTPUT_COLUMN_RENAME = {
    "Data_Operacao": "Data Operacao",
    "Dias_para_Encerramento": "Dias para Encerramento",
    "Dias_em_Operacao": "Dias em Operacao",
    "Ref_Entrada": "Ref Entrada",
    "Ref_Atual": "Ref Atual",
    "Soma_Bid_Offer": "Soma Bid Offer",
    "VARIACAO_ATIVO": "Variação Ativo",
    "RESULTADO_UNIT": "Resultado Unit",
    "RESULTADO_PCT": "Resultado %",
    "RESULTADO_TOTAL": "Resultado Total",
    "CDI_PCT": "CDI %",
    "Codigo_Produto": "Codigo Produto",
}

def _rename_output_headers(df: pd.DataFrame) -> pd.DataFrame:
    ren = {k: v for k, v in OUTPUT_COLUMN_RENAME.items() if k in df.columns}
    return df.rename(columns=ren)


# =========================
# Utilidades (texto/colunas)
# =========================

def _strip_accents(text: str) -> str:
    import unicodedata
    return "".join(c for c in unicodedata.normalize("NFD", text) if unicodedata.category(c) != "Mn")

def _norm_key(text: str) -> str:
    if text is None:
        return ""
    t = str(text)
    t = _strip_accents(t).lower()
    t = re.sub(r"[\s\-_\/\\]+", " ", t).strip()
    t = re.sub(r"[^a-z0-9 ]+", "", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t

def _colmap(df: pd.DataFrame) -> Dict[str, str]:
    return {_norm_key(c): c for c in df.columns}

def _resolve_column(df: pd.DataFrame, aliases: List[str], contains_tokens: Optional[List[List[str]]] = None) -> Optional[str]:
    cm = _colmap(df)
    # match exato por alias
    for a in aliases:
        k = _norm_key(a)
        if k in cm:
            return cm[k]
    # match por tokens
    if contains_tokens:
        keys = list(cm.keys())
        for token_list in contains_tokens:
            toks = [_norm_key(t) for t in token_list]
            for k in keys:
                if all(tok in k for tok in toks):
                    return cm[k]
    return None

def _to_numeric(s: pd.Series) -> pd.Series:
    """
    Converte Series para numérico preservando casas decimais.

    ⚠️ IMPORTANTE:
    - Quando o Excel é lido via openpyxl, colunas numéricas geralmente já vêm como float/int.
      Nesses casos NÃO devemos passar por .astype(str) nem remover '.', pois isso "anda" a vírgula
      (ex.: 0.022164... -> 2216483...).
    - Para valores em texto, trata os formatos mais comuns:
      * 1.234,56  (milhar com '.' e decimal com ',')
      * 1234,56   (decimal com ',')
      * 1234.56   (decimal com '.')
      * 1.234.567 (milhar com '.')
    """
    if s is None:
        return pd.Series(dtype="float64")

    # Já é numérico? mantém como numérico (preserva decimais)
    if pd.api.types.is_numeric_dtype(s):
        return pd.to_numeric(s, errors="coerce")

    x = s.astype(str).str.strip()

    # limpa espaços e símbolos comuns
    x = x.str.replace("\u00a0", "", regex=False)          # NBSP
    x = x.str.replace("%", "", regex=False)
    x = x.str.replace(r"[R$\s]", "", regex=True)

    # (123) -> -123
    x = x.str.replace("(", "-", regex=False).str.replace(")", "", regex=False)

    # Caso tenha '.' e ',' ao mesmo tempo => '.' milhar e ',' decimal
    mask_both = x.str.contains(r"\.") & x.str.contains(",")
    if mask_both.any():
        x.loc[mask_both] = x.loc[mask_both].str.replace(".", "", regex=False).str.replace(",", ".", regex=False)

    # Só vírgula => decimal com ','
    mask_comma = (~mask_both) & x.str.contains(",") & (~x.str.contains(r"\."))
    if mask_comma.any():
        x.loc[mask_comma] = x.loc[mask_comma].str.replace(",", ".", regex=False)

    # Só ponto: se for padrão de milhar (1.234.567) remove pontos, senão mantém como decimal
    mask_dot = (~mask_both) & x.str.contains(r"\.") & (~x.str.contains(","))
    if mask_dot.any():
        th_pat = r"^\-?\d{1,3}(\.\d{3})+$"
        mask_th = mask_dot & x.str.match(th_pat)
        if mask_th.any():
            x.loc[mask_th] = x.loc[mask_th].str.replace(".", "", regex=False)

    x = x.replace({"nan": np.nan, "None": np.nan, "": np.nan})
    return pd.to_numeric(x, errors="coerce")

def _normalize_conta_series(s: pd.Series) -> pd.Series:
    """Normaliza a coluna Conta para bater merges (remove .0 de floats, espaços e NBSP)."""
    if s is None:
        return pd.Series(dtype="object")
    x = s.astype(str).str.strip()
    x = x.str.replace("\u00a0", "", regex=False)
    # quando vem de float (ex.: 12345.0), remove o sufixo .0
    x = x.str.replace(r"\.0+$", "", regex=True)
    x = x.str.replace(r",0+$", "", regex=True)
    # remove 'nan'/'None'
    x = x.replace({"nan": np.nan, "None": np.nan, "": np.nan})
    return x





def _conta_key(s: pd.Series) -> pd.Series:
    """Chave canônica de Conta para merges (remove não-dígitos e zeros à esquerda)."""
    if s is None:
        return pd.Series(dtype="object")
    x = _normalize_conta_series(s).copy()
    # evita 'nan' como string
    x = x.fillna("")
    x = x.astype(str)
    x = x.str.replace(r"\D+", "", regex=True)     # só dígitos
    x = x.str.lstrip("0")                          # remove zeros à esquerda
    x = x.replace({"": np.nan})
    return x

def _parse_date_series(s: pd.Series) -> pd.Series:
    if s is None:
        return pd.Series([pd.NaT] * 0)
    # tenta dd/mm/yyyy, depois padrão pandas
    out = pd.to_datetime(s, errors="coerce", dayfirst=True)
    return out

def _safe_str(x) -> str:
    if pd.isna(x):
        return ""
    return str(x).strip()

def _first_non_null(x: pd.Series):
    x = x.dropna()
    return x.iloc[0] if not x.empty else np.nan

# =========================
# Desktop / busca de arquivos
# =========================

def _get_desktop_knownfolder() -> Optional[Path]:
    """
    Descobre o Desktop real do Windows (mesmo quando o usuário usa OneDrive / idioma diferente).
    """
    try:
        import ctypes
        from ctypes import wintypes

        # SHGetKnownFolderPath
        _SHGetKnownFolderPath = ctypes.windll.shell32.SHGetKnownFolderPath
        _SHGetKnownFolderPath.argtypes = [ctypes.c_void_p, wintypes.DWORD, wintypes.HANDLE, ctypes.POINTER(ctypes.c_wchar_p)]
        _SHGetKnownFolderPath.restype = wintypes.HRESULT

        # FOLDERID_Desktop
        import uuid
        FOLDERID_Desktop = uuid.UUID("{B4BFCC3A-DB2C-424C-B029-7FE99A87C641}")

        ppszPath = ctypes.c_wchar_p()
        hr = _SHGetKnownFolderPath(FOLDERID_Desktop.bytes_le, 0, 0, ctypes.byref(ppszPath))
        if hr != 0:
            return None
        return Path(ppszPath.value)
    except Exception:
        return None

def _candidate_search_dirs() -> List[Path]:
    home = Path.home()
    desktop = _get_desktop_knownfolder() or (home / "Desktop")
    # Desktop do OneDrive em PT/EN (tentativas)
    candidates = [desktop]
    # Planilhas base (pasta padrão das automações)
    try:
        if 'PLANILHAS_DIR' in globals() and PLANILHAS_DIR.exists():
            candidates.append(PLANILHAS_DIR)
    except Exception:
        pass
    # Downloads
    candidates.append(home / "Downloads")
    # Remove duplicados/ausentes
    out = []
    seen = set()
    for d in candidates:
        try:
            d = d.resolve()
        except Exception:
            pass
        if str(d).lower() in seen:
            continue
        seen.add(str(d).lower())
        if d.exists():
            out.append(d)
    return out

def _matches_patterns(filename: str, patterns: List[str]) -> bool:
    n = _norm_key(Path(filename).stem)
    for p in patterns:
        if _norm_key(p) in n:
            return True
    return False

def _find_latest_across_dirs(search_dirs: List[Path], patterns: List[str]) -> Optional[Path]:
    best: Optional[Tuple[float, Path]] = None
    for d in search_dirs:
        try:
            for ext in ("*.xlsx", "*.xlsm"):
                for f in d.glob(ext):
                    if _matches_patterns(f.name, patterns):
                        mtime = f.stat().st_mtime
                        if best is None or mtime > best[0]:
                            best = (mtime, f)
        except Exception:
            continue
    return best[1] if best else None

# =========================
# Leitura Excel (com fallback para arquivo travado)
# =========================

def _save_copy_via_excel_com(src: Path) -> Optional[Path]:
    """
    Se o arquivo estiver aberto no Excel e travado, tenta gerar uma cópia temporária usando COM.
    Retorna o caminho da cópia ou None.
    """
    try:
        import win32com.client  # type: ignore
    except Exception:
        return None

    tmpdir = Path(tempfile.mkdtemp(prefix="antecipacao_copy_"))
    dst = tmpdir / src.name

    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(str(src), ReadOnly=True)
        wb.SaveCopyAs(str(dst))
        wb.Close(False)
        excel.Quit()
        return dst
    except Exception:
        try:
            # tenta fechar Excel mesmo assim
            excel.Quit()
        except Exception:
            pass
        return None

def read_excel_safe(path: Path) -> pd.DataFrame:
    """
    Lê a primeira aba do Excel.
    - Se PermissionError: tenta salvar cópia via Excel COM (se disponível) e ler a cópia.
    """
    try:
        return pd.read_excel(path, engine="openpyxl")
    except PermissionError:
        tmp = _save_copy_via_excel_com(path)
        if tmp and tmp.exists():
            return pd.read_excel(tmp, engine="openpyxl")
        raise
    except Exception:
        # tenta novamente (OneDrive às vezes dá race condition)
        time.sleep(0.5)
        return pd.read_excel(path, engine="openpyxl")

# =========================
# Normalização e chaves
# =========================

@dataclass
class MonitorCols:
    conta: Optional[str]
    ativo: Optional[str]
    estrutura: Optional[str]
    data_operacao: Optional[str]
    fixing: Optional[str]
    quantidade: Optional[str]
    bid_offer: Optional[str]
    ref_atual: Optional[str]
    codigo_produto: Optional[str]
    barreira_knockin: Optional[str]
    barreira_knockout: Optional[str]
    knockin_hit: Optional[str]
    strike: Optional[str]
    tipo_operacao: Optional[str]
    tipo_opcao: Optional[str]
    cliente: Optional[str]

@dataclass
class CustodiaCols:
    conta: Optional[str]
    cliente: Optional[str]
    ref_entrada: Optional[str]
    ref_atual: Optional[str]
    quantidade: Optional[str]
    ativo: Optional[str]
    estrutura: Optional[str]
    data_operacao: Optional[str]
    fixing: Optional[str]

@dataclass
class AdvisorBaseCols:
    conta: Optional[str]
    assessor: Optional[str]
    cliente: Optional[str]

def detect_monitor_cols(df: pd.DataFrame) -> MonitorCols:
    return MonitorCols(
        conta=_resolve_column(df, ["Conta_Cliente", "Conta Cliente", "Conta"], contains_tokens=[["conta"]]),
        ativo=_resolve_column(df, ["Ativo", "Ticker", "Underlying"], contains_tokens=[["ativo"], ["ticker"], ["underlying"]]),
        estrutura=_resolve_column(df, ["Estrutura", "Strategy", "Produto"], contains_tokens=[["estrutura"], ["strategy"]]),
        data_operacao=_resolve_column(df, ["Data_Operação", "Data Operacao", "Trade Date"], contains_tokens=[["data", "operacao"], ["trade", "date"]]),
        fixing=_resolve_column(df, ["Fixing", "Vencimento", "Data Vencimento", "Maturity"], contains_tokens=[["fixing"], ["venc"], ["maturity"]]),
        quantidade=_resolve_column(df, ["Quantidade", "Qtd", "Quantity"], contains_tokens=[["quant"], ["qtd"], ["quantity"]]),
        bid_offer=_resolve_column(df, ["Bid(+)/Offer(-)", "Bid Offer", "Bid/Offer"], contains_tokens=[["bid"], ["offer"], ["bid", "offer"]]),
        ref_atual=_resolve_column(df, ["Ref", "Spot", "Preço Atual", "Ref Atual"], contains_tokens=[["ref"], ["spot"], ["preco", "atual"]]),
        codigo_produto=_resolve_column(df, ["Código do Produto", "Codigo do Produto", "Product Code"], contains_tokens=[["codigo", "produto"], ["product", "code"]]),
        barreira_knockin=_resolve_column(df, ["Barreira Knock In", "Barreira KnockIn", "Knock In"], contains_tokens=[["barreira", "knock", "in"], ["knock", "in"]]),
        barreira_knockout=_resolve_column(df, ["Barreira Knock Out", "Barreira KnockOut", "Knock Out"], contains_tokens=[["barreira", "knock", "out"], ["knock", "out"]]),
        knockin_hit=_resolve_column(df, ["KnockInAtingido", "Knock In Atingido", "KnockIn Hit"], contains_tokens=[["knockin"], ["atingido"], ["hit"]]),
        strike=_resolve_column(df, ["Preço Exercício", "Preco Exercicio", "Strike"], contains_tokens=[["preco", "exerc"], ["strike"]]),
        tipo_operacao=_resolve_column(df, ["Tipo Operação", "Tipo Operacao", "Side"], contains_tokens=[["tipo", "oper"], ["side"]]),
        tipo_opcao=_resolve_column(df, ["Tipo Opção", "Tipo Opcao", "Option Type"], contains_tokens=[["tipo", "op"], ["option", "type"]]),
        cliente=_resolve_column(df, ["Cliente", "Nome Cliente"], contains_tokens=[["nome", "cliente"]]),
    )

def detect_custodia_cols(df: pd.DataFrame) -> CustodiaCols:
    return CustodiaCols(
        conta=_resolve_column(df, ["Conta", "Conta_Cliente", "Conta Cliente"], contains_tokens=[["conta"]]),
        cliente=_resolve_column(df, ["Cliente", "Nome Cliente"], contains_tokens=[["nome", "cliente"]]),
        ref_entrada=_resolve_column(df, ["Ref entrada", "Ref_Entrada", "Preço Inicial de Referência", "Preco Inicial de Referencia", "Preço de Referência de Entrada"], contains_tokens=[["ref", "entrada"], ["preco", "inicial"], ["referencia", "entrada"]]),
        ref_atual=_resolve_column(df, ["Ref", "Ref atual", "Preço Atual", "Spot"], contains_tokens=[["ref"], ["preco", "atual"], ["spot"]]),
        quantidade=_resolve_column(df, ["Quantidade", "Qtd", "Quantity"], contains_tokens=[["quant"], ["qtd"], ["quantity"]]),
        ativo=_resolve_column(df, ["Ativo", "Ticker", "Underlying"], contains_tokens=[["ativo"], ["ticker"], ["underlying"]]),
        estrutura=_resolve_column(df, ["Estrutura", "Strategy", "Produto"], contains_tokens=[["estrutura"], ["strategy"]]),
        data_operacao=_resolve_column(df, ["Data_Operação", "Data Operacao", "Trade Date"], contains_tokens=[["data", "operacao"], ["trade", "date"]]),
        fixing=_resolve_column(df, ["Fixing", "Vencimento", "Data Vencimento", "Maturity"], contains_tokens=[["fixing"], ["venc"], ["maturity"]]),
    )

def detect_advisor_base_cols(df: pd.DataFrame) -> AdvisorBaseCols:
    return AdvisorBaseCols(
        conta=_resolve_column(df, ["Conta", "Conta_Cliente", "Conta Cliente"], contains_tokens=[["conta"]]),
        assessor=_resolve_column(df, ["Assessor", "Advisor", "AAI", "Assessoria"], contains_tokens=[["assessor"], ["advisor"], ["aai"]]),
        cliente=_resolve_column(df, ["Cliente", "Nome Cliente"], contains_tokens=[["nome", "cliente"]]),
    )

def build_op_key(df: pd.DataFrame, cols: List[str], fallback_col: Optional[str] = None) -> pd.Series:
    """
    Cria uma chave de operação unindo colunas disponíveis.
    """
    use_cols = [c for c in cols if c and c in df.columns]
    if not use_cols and fallback_col and fallback_col in df.columns:
        return df[fallback_col].astype(str).fillna("").map(lambda x: _norm_key(x))
    if not use_cols:
        return pd.Series([str(i) for i in range(len(df))], index=df.index)

    parts = []
    for c in use_cols:
        if np.issubdtype(df[c].dtype, np.datetime64):
            parts.append(df[c].dt.strftime("%Y-%m-%d").fillna(""))
        else:
            parts.append(df[c].astype(str).fillna("").map(lambda x: _norm_key(x)))
    key = parts[0]
    for p in parts[1:]:
        key = key + "|" + p
    return key

# =========================
# CDI (BCB/SGS) – cache e cálculo acumulado
# =========================

def _fetch_cdi_series(start: _dt.date, end: _dt.date) -> pd.DataFrame:
    """
    Busca CDI diário (% ao dia) via API do BCB/SGS.

    Endpoint padrão do BCData/SGS:
    https://api.bcb.gov.br/dados/serie/bcdata.sgs.{codigo}/dados?formato=json&dataInicial=dd/MM/aaaa&dataFinal=dd/MM/aaaa
    """
    import requests  # requests é padrão em muitos ambientes; se não tiver, o erro será exibido pro usuário.

    url = f"https://api.bcb.gov.br/dados/serie/bcdata.sgs.{BCB_SGS_SERIE_CDI_DIA}/dados"
    params = {
        "formato": "json",
        "dataInicial": start.strftime("%d/%m/%Y"),
        "dataFinal": end.strftime("%d/%m/%Y"),
    }
    r = requests.get(url, params=params, timeout=BCB_TIMEOUT_SEC)
    r.raise_for_status()
    data = r.json()

    df = pd.DataFrame(data)
    if df.empty:
        return df
    # API traz "data" e "valor"
    df["data"] = pd.to_datetime(df["data"], format="%d/%m/%Y", errors="coerce")
    df["valor"] = pd.to_numeric(df["valor"].astype(str).str.replace(",", ".", regex=False), errors="coerce")
    df = df.dropna(subset=["data", "valor"]).sort_values("data")
    return df

def build_cdi_prefix(min_date: _dt.date, max_date: _dt.date) -> Tuple[pd.DatetimeIndex, np.ndarray]:
    """
    Retorna (datas, cum_factor) onde cum_factor[i] = prod(1 + CDI_dia/100) até a data i.
    """
    df = _fetch_cdi_series(min_date, max_date)
    if df.empty:
        return pd.DatetimeIndex([]), np.array([], dtype="float64")
    factors = 1.0 + (df["valor"].values / 100.0)
    cum = np.cumprod(factors)
    return pd.DatetimeIndex(df["data"]), cum

def cdi_period_pct(start_dt: pd.Timestamp, idx: pd.DatetimeIndex, cum: np.ndarray) -> float:
    """
    CDI acumulado do período [start_dt, end] em fração (0.10 = 10%).
    """
    if idx.empty or cum.size == 0 or pd.isna(start_dt):
        return np.nan
    start_dt = pd.to_datetime(start_dt).normalize()
    # primeiro índice com data >= start_dt
    pos = idx.searchsorted(start_dt, side="left")
    if pos >= len(idx):
        return np.nan
    end_factor = cum[-1]
    denom = cum[pos - 1] if pos > 0 else 1.0
    return (end_factor / denom) - 1.0

# =========================
# Construção da base (Monitor + Custódia + Advisor Base)
# =========================

def build_monitor_operations(df_mon: pd.DataFrame) -> Tuple[pd.DataFrame, MonitorCols]:
    mc = detect_monitor_cols(df_mon)
    df = df_mon.copy()

    # normaliza numéricos importantes
    if mc.bid_offer and mc.bid_offer in df.columns:
        df[mc.bid_offer] = _to_numeric(df[mc.bid_offer])
    if mc.ref_atual and mc.ref_atual in df.columns:
        df[mc.ref_atual] = _to_numeric(df[mc.ref_atual])
    if mc.quantidade and mc.quantidade in df.columns:
        df[mc.quantidade] = _to_numeric(df[mc.quantidade])

    # normaliza datas
    for dc in [mc.data_operacao, mc.fixing]:
        if dc and dc in df.columns:
            df[dc] = _parse_date_series(df[dc])

    # chave de operação
    # Regra prática: se existir "Código do Produto" (e Conta), usamos Conta+Código como chave.
    # Isso garante que as pernas do monitor agrupem corretamente e que o merge com a custódia bata 1:1.
    if mc.codigo_produto and mc.codigo_produto in df.columns and mc.conta and mc.conta in df.columns:
        op_cols = [mc.conta, mc.codigo_produto]
        df["_op_key"] = build_op_key(df, op_cols, fallback_col=None)
    else:
        # fallback (espírito do modelo antigo): agrupa por campos de operação
        op_cols = [mc.conta, mc.ativo, mc.estrutura, mc.data_operacao, mc.fixing, mc.quantidade]
        # se o monitor ainda tiver ref entrada, inclui (quando existir)
        ref_ent_mon = _resolve_column(df, ["Ref entrada", "Ref_Entrada"], contains_tokens=[["ref", "entrada"]])
        if ref_ent_mon and ref_ent_mon in df.columns:
            df[ref_ent_mon] = _to_numeric(df[ref_ent_mon])
            op_cols.append(ref_ent_mon)
        df["_op_key"] = build_op_key(df, op_cols, fallback_col=mc.codigo_produto)

    # Soma_Bid_Offer por operação (IGUAL ao modelo antigo)
    if mc.bid_offer and mc.bid_offer in df.columns:
        df["Soma_Bid_Offer"] = df.groupby("_op_key", dropna=False)[mc.bid_offer].transform("sum")
    else:
        df["Soma_Bid_Offer"] = np.nan

    # Preço de Ex Knock In (igual modelo antigo: BUY + CALL => Strike)
    preco_ex = pd.Series([np.nan] * len(df), index=df.index, dtype="float64")
    if mc.strike and mc.tipo_operacao and mc.tipo_opcao:
        if mc.strike in df.columns and mc.tipo_operacao in df.columns and mc.tipo_opcao in df.columns:
            strike = _to_numeric(df[mc.strike])
            side = df[mc.tipo_operacao].astype(str).str.upper()
            opt = df[mc.tipo_opcao].astype(str).str.upper()
            mask = (side == "BUY") & (opt == "CALL")
            preco_ex = np.where(mask, strike, np.nan)
    df["Preço de Ex Knock In"] = preco_ex
    # reduz para 1 linha por operação (mantendo Soma_Bid_Offer já calculada)
    agg = {}
    # campos base para o 1º não nulo
    keep_first = [
        ("Conta", mc.conta),
        ("Ativo", mc.ativo),
        ("Estrutura", mc.estrutura),
        ("Data_Operacao", mc.data_operacao),
        ("Fixing", mc.fixing),
        ("Quantidade_mon", mc.quantidade),
        ("Ref_Atual_mon", mc.ref_atual),
        ("Cliente_mon", mc.cliente),
        ("Codigo_Produto_mon", mc.codigo_produto),
        ("Barreira Knock In", mc.barreira_knockin),
        ("Barreira Knock Out", mc.barreira_knockout),
        ("KnockInAtingido", mc.knockin_hit),
    ]
    for out, src in keep_first:
        if src and src in df.columns:
            if src in [mc.data_operacao]:
                agg[src] = "min"
            elif src in [mc.fixing]:
                agg[src] = "max"
            else:
                agg[src] = _first_non_null

    # Soma_Bid_Offer precisa ser a soma => já está repetido em todas as linhas do grupo, então first serve
    agg["Soma_Bid_Offer"] = _first_non_null
    agg["Preço de Ex Knock In"] = _first_non_null

    df_op = df.groupby("_op_key", dropna=False).agg(agg).reset_index()

    # renomeia para nomes padrão
    ren = {}
    if mc.conta and mc.conta in df_op.columns: ren[mc.conta] = "Conta"
    if mc.ativo and mc.ativo in df_op.columns: ren[mc.ativo] = "Ativo"
    if mc.estrutura and mc.estrutura in df_op.columns: ren[mc.estrutura] = "Estrutura"
    if mc.data_operacao and mc.data_operacao in df_op.columns: ren[mc.data_operacao] = "Data_Operacao"
    if mc.fixing and mc.fixing in df_op.columns: ren[mc.fixing] = "Fixing"
    if mc.quantidade and mc.quantidade in df_op.columns: ren[mc.quantidade] = "Quantidade"
    if mc.ref_atual and mc.ref_atual in df_op.columns: ren[mc.ref_atual] = "Ref_Atual"
    if mc.cliente and mc.cliente in df_op.columns: ren[mc.cliente] = "Cliente"
    if mc.codigo_produto and mc.codigo_produto in df_op.columns: ren[mc.codigo_produto] = "Codigo_Produto"
    df_op = df_op.rename(columns=ren)

    # fallback de quantidade/ref_atual se vierem dos aliases
    if "Quantidade_mon" in df_op.columns and "Quantidade" not in df_op.columns:
        df_op["Quantidade"] = df_op["Quantidade_mon"]
    if "Ref_Atual_mon" in df_op.columns and "Ref_Atual" not in df_op.columns:
        df_op["Ref_Atual"] = df_op["Ref_Atual_mon"]
    if "Cliente_mon" in df_op.columns and "Cliente" not in df_op.columns:
        df_op["Cliente"] = df_op["Cliente_mon"]
    if "Codigo_Produto_mon" in df_op.columns and "Codigo_Produto" not in df_op.columns:
        df_op["Codigo_Produto"] = df_op["Codigo_Produto_mon"]

    return df_op, mc

def build_custodia_head(df_cust: pd.DataFrame) -> Tuple[pd.DataFrame, CustodiaCols]:
    cc = detect_custodia_cols(df_cust)
    df = df_cust.copy()

    # normaliza numéricos
    for col in [cc.ref_entrada, cc.ref_atual, cc.quantidade]:
        if col and col in df.columns:
            df[col] = _to_numeric(df[col])

    # normaliza datas
    for dc in [cc.data_operacao, cc.fixing]:
        if dc and dc in df.columns:
            df[dc] = _parse_date_series(df[dc])

    # chave de operação (compatível com monitor)
    # Preferimos Conta + Código do Produto quando disponível (merge robusto).
    codigo_prod = _resolve_column(df, ["Código do Produto", "Codigo do Produto", "Product Code"], contains_tokens=[["codigo", "produto"], ["product", "code"]])
    if codigo_prod and codigo_prod in df.columns and cc.conta and cc.conta in df.columns:
        op_cols = [cc.conta, codigo_prod]
        df["_op_key"] = build_op_key(df, op_cols, fallback_col=None)
    else:
        # fallback: NÃO inclui Ref_Entrada na chave (senão o merge com monitor quebra)
        op_cols = [cc.conta, cc.ativo, cc.estrutura, cc.data_operacao, cc.fixing, cc.quantidade]
        df["_op_key"] = build_op_key(df, op_cols, fallback_col=codigo_prod)

    # reduz para 1 linha por operação
    agg = {}
    for k in [cc.conta, cc.cliente, cc.ref_entrada, cc.ref_atual, cc.quantidade, cc.ativo, cc.estrutura, cc.data_operacao, cc.fixing]:
        if k and k in df.columns:
            if k == cc.data_operacao:
                agg[k] = "min"
            elif k == cc.fixing:
                agg[k] = "max"
            else:
                agg[k] = _first_non_null

    head = df.groupby("_op_key", dropna=False).agg(agg).reset_index()

    # renomeia para padrão
    ren = {}
    if cc.conta and cc.conta in head.columns: ren[cc.conta] = "Conta"
    if cc.cliente and cc.cliente in head.columns: ren[cc.cliente] = "Cliente"
    if cc.ref_entrada and cc.ref_entrada in head.columns: ren[cc.ref_entrada] = "Ref_Entrada"
    if cc.ref_atual and cc.ref_atual in head.columns: ren[cc.ref_atual] = "Ref_Atual_cust"
    if cc.quantidade and cc.quantidade in head.columns: ren[cc.quantidade] = "Quantidade_cust"
    if cc.ativo and cc.ativo in head.columns and "Ativo" not in head.columns: ren[cc.ativo] = "Ativo"
    if cc.estrutura and cc.estrutura in head.columns and "Estrutura" not in head.columns: ren[cc.estrutura] = "Estrutura"
    if cc.data_operacao and cc.data_operacao in head.columns and "Data_Operacao" not in head.columns: ren[cc.data_operacao] = "Data_Operacao"
    if cc.fixing and cc.fixing in head.columns and "Fixing" not in head.columns: ren[cc.fixing] = "Fixing"
    head = head.rename(columns=ren)

    return head, cc

def load_advisor_base(path: Path) -> Optional[pd.DataFrame]:
    try:
        df = pd.read_excel(path, engine="openpyxl")
        bc = detect_advisor_base_cols(df)
        if not bc.conta or not bc.assessor:
            return None
        out = pd.DataFrame({
            "Conta": _normalize_conta_series(df[bc.conta]),
            "Assessor": df[bc.assessor].astype(str).str.strip()
        })
        if bc.cliente and bc.cliente in df.columns:
            out["Cliente_base"] = df[bc.cliente].astype(str).str.strip()
        return out.dropna(subset=["Conta"])
    except Exception:
        return None

# =========================
# Cálculo principal
# =========================

def calcular_antecipacao(monitor_path: Optional[Path], custodia_path: Optional[Path], advisor_base_path: Optional[Path]) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    hoje = pd.Timestamp(_dt.date.today())

    df_mon = read_excel_safe(monitor_path) if monitor_path else None
    df_cust = read_excel_safe(custodia_path) if custodia_path else None

    mon_op = None
    cust_head = None

    if df_mon is not None:
        mon_op, _mc = build_monitor_operations(df_mon)

    if df_cust is not None:
        cust_head, _cc = build_custodia_head(df_cust)

    if mon_op is None and cust_head is None:
        raise FileNotFoundError("Não encontrei nenhuma planilha para processar (Monitor e/ou Custódia).")

    # base inicial
    if mon_op is not None:
        base = mon_op.copy()
    else:
        base = cust_head.copy()

    # enriquece com custódia
    if cust_head is not None:
        base = base.merge(cust_head, on="_op_key", how="left", suffixes=("", "_cust"))
        # Ref_Entrada vem da custódia (regra do projeto)
        if "Ref_Entrada" not in base.columns:
            base["Ref_Entrada"] = np.nan

        # Ref_Atual: prioriza monitor, senão custódia
        if "Ref_Atual" not in base.columns or base["Ref_Atual"].isna().all():
            if "Ref_Atual_cust" in base.columns:
                base["Ref_Atual"] = base.get("Ref_Atual", pd.Series([np.nan]*len(base))).fillna(base["Ref_Atual_cust"])

        # Quantidade: prioriza custódia, senão monitor
        qtd = pd.Series([np.nan] * len(base))
        if "Quantidade_cust" in base.columns:
            qtd = base["Quantidade_cust"]
        if "Quantidade" in base.columns:
            qtd = qtd.fillna(base["Quantidade"])
        base["Quantidade"] = qtd

        # Cliente: prioriza custódia, senão monitor
        if "Cliente" in base.columns:
            if "Cliente_mon" in base.columns:
                base["Cliente"] = base["Cliente"].fillna(base["Cliente_mon"])
        elif "Cliente_mon" in base.columns:
            base["Cliente"] = base["Cliente_mon"]
        else:
            base["Cliente"] = np.nan

    # se só custódia, garante campos esperados
    for c in ["Conta", "Ativo", "Estrutura", "Data_Operacao", "Fixing", "Ref_Entrada", "Ref_Atual", "Quantidade", "Soma_Bid_Offer"]:
        if c not in base.columns:
            base[c] = np.nan

    # Datas e dias
    base["Data_Operacao_dt"] = pd.to_datetime(base["Data_Operacao"], errors="coerce")
    base["Fixing_dt"] = pd.to_datetime(base["Fixing"], errors="coerce")

    base["Dias_para_Encerramento"] = (base["Fixing_dt"] - hoje).dt.days
    base["Dias_em_Operacao"] = (hoje - base["Data_Operacao_dt"]).dt.days

    # Cálculos
    base["VARIACAO_ATIVO"] = base["Ref_Atual"] / base["Ref_Entrada"] - 1
    base["RESULTADO_UNIT"] = base["Soma_Bid_Offer"] + (base["Ref_Atual"] - base["Ref_Entrada"])
    base["RESULTADO_PCT"] = base["RESULTADO_UNIT"] / base["Ref_Entrada"]
    base["RESULTADO_TOTAL"] = base["RESULTADO_UNIT"] * base["Quantidade"]

    # =========================
    # CDI do período (Data_Operacao -> hoje)
    # =========================
    valid_dates = base["Data_Operacao_dt"].dropna()
    if not valid_dates.empty:
        min_date = valid_dates.min().date()
        max_date = hoje.date()
        try:
            idx, cum = build_cdi_prefix(min_date, max_date)
            base["CDI_PCT"] = base["Data_Operacao_dt"].apply(lambda d: cdi_period_pct(d, idx, cum))
        except Exception as e:
            print(f"⚠️ Não consegui calcular CDI via BCB (seguindo sem CDI): {e}")
            base["CDI_PCT"] = np.nan
    else:
        base["CDI_PCT"] = np.nan

    # =========================
    # Advisor Base (Assessor)
    # =========================
    if advisor_base_path and advisor_base_path.exists():
        df_base = load_advisor_base(advisor_base_path)
        if df_base is not None:
            # Merge por chave canônica (resolve casos com zeros à esquerda na base vs. Conta numérica no monitor)
            base["_ContaKey"] = _conta_key(base["Conta"])
            df_base["_ContaKey"] = _conta_key(df_base["Conta"])
            base = base.merge(df_base.drop(columns=["Conta"], errors="ignore"), on="_ContaKey", how="left")
            base.drop(columns=["_ContaKey"], inplace=True, errors="ignore")
            # se Cliente vazio, tenta preencher da base
            if "Cliente_base" in base.columns:
                base["Cliente"] = base["Cliente"].fillna(base["Cliente_base"])
    if "Assessor" not in base.columns:
        base["Assessor"] = np.nan

    # Datas para exibição
    base["Data_Operacao"] = base["Data_Operacao_dt"].dt.strftime("%d/%b/%Y")
    base["Fixing"] = base["Fixing_dt"].dt.strftime("%d/%b/%Y")

    # =========================
    # Abas de saída
    # =========================
    colunas = [
        "Conta", "Cliente", "Assessor",
        "Data_Operacao", "Fixing",
        "Dias_para_Encerramento", "Dias_em_Operacao",
        "Ativo", "Estrutura",
        "Ref_Entrada", "Ref_Atual", "Quantidade",
        "Soma_Bid_Offer", "VARIACAO_ATIVO",
        "RESULTADO_UNIT", "RESULTADO_PCT", "RESULTADO_TOTAL",
        "CDI_PCT",
        # opcionais
        "Codigo_Produto", "Barreira Knock In", "Barreira Knock Out", "KnockInAtingido", "Preço de Ex Knock In",
    ]
    # mantém só o que existe
    colunas_exist = [c for c in colunas if c in base.columns]

    dados_crus = base[colunas_exist].copy()

    # ganho = resultado % > 0 (se quiser mudar, aqui é o lugar)
    df_ganho = base[base["RESULTADO_PCT"] > 0].copy()
    df_geral = base.copy()

    # antec com ganho – remove colunas solicitadas
    drop_ganho = ["Codigo_Produto", "Barreira Knock In", "Barreira Knock Out", "KnockInAtingido", "Preço de Ex Knock In"]
    for c in drop_ganho:
        if c in df_ganho.columns:
            df_ganho.drop(columns=[c], inplace=True)

    # reordena
    df_ganho = df_ganho[[c for c in colunas_exist if c in df_ganho.columns]]
    df_geral = df_geral[colunas_exist]

    return dados_crus, df_ganho, df_geral

# =========================
# Excel: salvar e formatar
# =========================

from openpyxl.utils import get_column_letter

def salvar_relatorio(out_path: Path, dados_crus: pd.DataFrame, df_ganho: pd.DataFrame, df_geral: pd.DataFrame) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        dados_crus.to_excel(writer, sheet_name="Dados Crus", index=False)
        df_ganho.to_excel(writer, sheet_name="Antecipacao com Ganho", index=False)
        df_geral.to_excel(writer, sheet_name="Resultado Geral", index=False)

    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, numbers

    wb = load_workbook(out_path)
    header_fill = PatternFill(start_color="0D005C", end_color="0D005C", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    cell_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    border_style = Side(border_style="thin", color="FFFFFF")
    
    # ALTERAÇÃO: wrap_text alterado para False
    alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=False)

    for ws in wb.worksheets:
        # cabeçalho
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
            cell.alignment = alignment_center

        # corpo
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.fill = cell_fill
                cell.border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
                cell.alignment = alignment_center

        # Formatos numéricos (mantidos conforme seu código)
        header = [c.value for c in ws[1]]
        def _col_idx(name: str):
            try: return header.index(name) + 1
            except ValueError: return None

        for name in ["RESULTADO_PCT", "VARIACAO_ATIVO", "CDI_PCT"]:
            ci = _col_idx(name)
            if ci:
                for cell in ws.iter_cols(min_col=ci, max_col=ci, min_row=2):
                    for c in cell: c.number_format = numbers.FORMAT_PERCENTAGE_00

        # RESULTADOS (formatação)
        # Obs: Soma_Bid_Offer e RESULTADO_UNIT podem ter várias casas decimais no Monitor.
        # Para evitar a impressão de "erro de soma" por arredondamento visual, deixamos 6 casas.
        fmt_money2 = '"R$" #,##0.00'
        fmt_money6 = '"R$" #,##0.000000'

        for name in ["RESULTADO_UNIT", "Soma_Bid_Offer"]:
            ci = _col_idx(name)
            if ci:
                for cell in ws.iter_cols(min_col=ci, max_col=ci, min_row=2):
                    for c in cell:
                        c.number_format = fmt_money6

        ci = _col_idx("RESULTADO_TOTAL")
        if ci:
            for cell in ws.iter_cols(min_col=ci, max_col=ci, min_row=2):
                for c in cell:
                    c.number_format = fmt_money2

        # NOVA PARTE: Ajuste automático de largura (AutoFit manual)
        for col in ws.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            
            for cell in col:
                try:
                    if cell.value:
                        val_len = len(str(cell.value))
                        if val_len > max_length:
                            max_length = val_len
                except:
                    pass
            
            # Define a largura com uma margem de segurança de 2 caracteres
            ws.column_dimensions[column_letter].width = max_length + 2

        ws.sheet_view.showGridLines = False

    wb.save(out_path)

# =========================
# Email: corpo com tabela Resultado% > CDI
# =========================

def _html_table(df: pd.DataFrame, cols: List[str]) -> str:
    """
    Tabela HTML com o padrão visual do relatório (Calibri 12 + header azul).
    """
    df2 = df.copy()
    df2 = df2[cols].copy()

    # formata percentuais
    for c in ["Resultado %", "CDI %", "Resultado_PCT", "CDI_PCT"]:
        if c in df2.columns:
            df2[c] = pd.to_numeric(df2[c], errors="coerce").map(lambda x: "" if pd.isna(x) else f"{x*100:,.2f}%".replace(",", "X").replace(".", ",").replace("X", "."))

    # style
    head_style = "background:#0D005C;color:#FFFFFF;font-weight:bold;border:1px solid #FFFFFF;padding:6px;text-align:center;"
    cell_style = "background:#C0C0C0;border:1px solid #FFFFFF;padding:6px;text-align:center;"
    table_style = "border-collapse:collapse;font-family:Calibri,Arial;font-size:12px;"

    html = [f'<table style="{table_style}">', "<thead><tr>"]
    for c in df2.columns:
        html.append(f'<th style="{head_style}">{c}</th>')
    html.append("</tr></thead><tbody>")
    for _, row in df2.iterrows():
        html.append("<tr>")
        for c in df2.columns:
            v = row[c]
            if pd.isna(v):
                v = ""
            html.append(f'<td style="{cell_style}">{v}</td>')
        html.append("</tr>")
    html.append("</tbody></table>")
    return "".join(html)

def tentar_enviar_email(out_path: Path, df_geral: pd.DataFrame, base_dir_for_signature: Path) -> None:
    if not ENVIAR_EMAIL:
        return
    try:
        import win32com.client as win32  # type: ignore
    except Exception as e:
        print(f"⚠️ pywin32 não disponível. Não enviei e-mail. ({e})")
        return

    # assinatura
    sig = None
    if ASSINATURA_FIXA and Path(ASSINATURA_FIXA).exists():
        sig = Path(ASSINATURA_FIXA)
    else:
        # procura no base_dir e no Desktop real
        for cand in [base_dir_for_signature, (_get_desktop_knownfolder() or Path.home() / "Desktop")]:
            try:
                f = next(cand.glob("Assinatura*.png"), None)
                if f and f.exists():
                    sig = f
                    break
            except Exception:
                continue

    # operações que batem CDI
    cols_email = ["Conta", "Cliente", "Assessor", "Ativo", "Estrutura", "Dias_para_Encerramento", "RESULTADO_PCT", "CDI_PCT"]
    tmp = df_geral.copy()
    for c in ["RESULTADO_PCT", "CDI_PCT"]:
        if c in tmp.columns:
            tmp[c] = pd.to_numeric(tmp[c], errors="coerce")
    tmp = tmp.dropna(subset=["RESULTADO_PCT", "CDI_PCT"])
    tmp = tmp[tmp["RESULTADO_PCT"] > tmp["CDI_PCT"]].copy()

    # monta bloco HTML (sempre aparece no e-mail)
    bloco_tabela = "<p><b>Operações com Resultado % acima do CDI do período (Data Operação → Hoje):</b></p>"
    if tmp.empty:
        bloco_tabela += "<p><i>Nenhuma operação superou o CDI do período nas informações disponíveis.</i></p>"
    else:
        tmp = tmp.sort_values("RESULTADO_PCT", ascending=False)
        tmp2 = tmp[cols_email].rename(columns={
            "Dias_para_Encerramento": "Dias para Encerramento",
            "RESULTADO_PCT": "Resultado %",
            "CDI_PCT": "CDI %",
        })
        show_cols = ["Conta", "Cliente", "Assessor", "Ativo", "Estrutura", "Dias para Encerramento", "Resultado %", "CDI %"]
        bloco_tabela += _html_table(tmp2, show_cols)
    outlook = win32.Dispatch("Outlook.Application")
    mail = outlook.CreateItem(0)

    body = f"""
    <html>
      <body style="font-family:Calibri,Arial;font-size:12px;">
        <p>Segue em anexo o relatório atualizado das operações para antecipação.</p>
        {bloco_tabela}
        <p><i>*Operações sem resultado geralmente indicam ausência do preço de referência de entrada.</i></p>
        <p>Atenciosamente,</p>
        {"<p><img src='cid:assinatura' width='400'></p>" if sig else ""}
      </body>
    </html>
    """

    mail.To = EMAIL_TO
    mail.Subject = EMAIL_SUBJECT
    mail.HTMLBody = body
    mail.Attachments.Add(str(out_path))

    if sig:
        att = mail.Attachments.Add(str(sig))
        # inline cid
        att.PropertyAccessor.SetProperty("http://schemas.microsoft.com/mapi/proptag/0x3712001F", "assinatura")

    mail.Send()
    print(f"📧 E-mail enviado para {EMAIL_TO}")

# =========================
# Main
# =========================

def choose_output_base(desktop_primary: Path, monitor: Optional[Path], custodia: Optional[Path], fallback_dir: Path) -> Path:
    # preferência: Planilhas base (se existir), depois mesma pasta onde achou os arquivos
    try:
        if 'PLANILHAS_DIR' in globals() and PLANILHAS_DIR.exists():
            return PLANILHAS_DIR
    except Exception:
        pass
    for p in [monitor, custodia]:
        if p and p.exists():
            return p.parent
    return desktop_primary if desktop_primary.exists() else fallback_dir

def main() -> None:
    desktop_primary = _get_desktop_knownfolder() or (Path.home() / "Desktop")
    search_dirs = _candidate_search_dirs()

    print(f"📌 Desktop (KnownFolder): {desktop_primary}")
    print("🔍 Pastas pesquisadas:")
    for d in search_dirs:
        print(f"  - {d}")

    # arquivos
    monitor = Path(FORCAR_ARQUIVO_MONITOR).expanduser() if FORCAR_ARQUIVO_MONITOR else _find_latest_across_dirs(search_dirs, PADROES_MONITOR)
    custodia = Path(FORCAR_ARQUIVO_CUSTODIA).expanduser() if FORCAR_ARQUIVO_CUSTODIA else _find_latest_across_dirs(search_dirs, PADROES_CUSTODIA)
    advisor_base = Path(FORCAR_ARQUIVO_ADVISOR_BASE).expanduser() if FORCAR_ARQUIVO_ADVISOR_BASE else _find_latest_across_dirs(search_dirs, PADROES_ADVISOR_BASE)

    print(f"📄 Monitor: {monitor if monitor else '(não encontrado)'}")
    print(f"📄 Custódia: {custodia if custodia else '(não encontrado)'}")
    print(f"📄 Advisor Base: {advisor_base if advisor_base else '(não encontrado)'}")

    dados_crus, df_ganho, df_geral = calcular_antecipacao(monitor, custodia, advisor_base)

    base_out = choose_output_base(desktop_primary, monitor, custodia, fallback_dir=Path(__file__).resolve().parent)
    out_dir = base_out / NOME_PASTA_SAIDA
    out_path = out_dir / NOME_ARQUIVO_SAIDA

    salvar_relatorio(out_path, dados_crus, df_ganho, df_geral)

    print(f"✅ Relatório gerado em: {out_path}")

    try:
        tentar_enviar_email(out_path, df_geral, base_out)
    except Exception as e:
        print(f"⚠️ Não consegui enviar e-mail (sem travar o processo): {e}")

if __name__ == "__main__":
    try:
        main()
    except SystemExit:
        raise
    except Exception:
        _log_unhandled_exception("run_err")
        raise
