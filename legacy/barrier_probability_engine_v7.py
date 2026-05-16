# -*- coding: utf-8 -*-

# PUBLIC-SAFE PORTFOLIO VERSION
# Real credentials, client data, private endpoints and internal paths were removed or replaced by placeholders.
"""
calcula_prob_barreira_v7_noCOM.py
Modo "IDLE-first": NÃO usa Excel COM (não abre/fecha Excel).
Fluxo:
  1) 1ª execução: lê Operações - DATA + SI_D_SEDE -> gera RTD_TICKERS (lista de opções a exportar no Profit). Encerra.
  2) Você exporta do Profit para a aba RTD_EXPORT (RTD/DDE) e SALVA a planilha.
  3) 2ª execução: lê RTD_EXPORT -> calcula smile/skew simples -> σ_barreira -> P(tocar barreira por FECHAMENTO) e grava RESULTADOS.
Requisitos: openpyxl, numpy, requests (opcional p/ juros SGS).
"""

from __future__ import annotations
import os
import re
import math
import sys
import json
import time
import argparse
import datetime as dt
from pathlib import Path
from typing import Dict, List, Tuple, Optional

import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

try:
    import requests
except Exception:
    requests = None


# --------------------------
# Helpers
# --------------------------

def _norm(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"\s+", " ", s)
    s = s.replace("ç", "c").replace("ã", "a").replace("á", "a").replace("à", "a").replace("â", "a")
    s = s.replace("é", "e").replace("ê", "e").replace("í", "i").replace("ó", "o").replace("ô", "o").replace("õ", "o")
    s = s.replace("ú", "u")
    return s

def _to_float(x):
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip()
    if s == "":
        return None
    # percent with comma
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None

def _to_date(x):
    if x is None:
        return None
    if isinstance(x, dt.datetime):
        return x.date()
    if isinstance(x, dt.date):
        return x
    s = str(x).strip()
    # dd/mm/yyyy
    m = re.match(r"^(\d{2})/(\d{2})/(\d{4})$", s)
    if m:
        d, mo, y = map(int, m.groups())
        return dt.date(y, mo, d)
    # yyyy-mm-dd
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", s)
    if m:
        y, mo, d = map(int, m.groups())
        return dt.date(y, mo, d)
    return None

def _pct_to_frac(p):
    """Accepts 0.90, 90, '90,00%' etc -> 0.90"""
    if p is None:
        return None
    if isinstance(p, str) and "%" in p:
        p = p.replace("%", "")
    v = _to_float(p)
    if v is None:
        return None
    if v > 2.0:  # 90 -> 0.90
        v = v / 100.0
    return v

def _first4_alnum(s: str) -> str:
    s = re.sub(r"[^A-Za-z0-9]", "", (s or ""))
    return s[:4].upper()

def _safe_sheet(wb, name: str):
    if name in wb.sheetnames:
        return wb[name]
    return wb.create_sheet(name)

def _clear_sheet(ws):
    ws.delete_rows(1, ws.max_row or 1)

def _sgs_daily_rate(sgs_code: int = 11) -> Optional[float]:
    """
    Busca última taxa diária via BCB/SGS (retorna anualizada aproximada em decimal).
    Série 11 = SELIC diária (% a.a.)
    """
    if requests is None:
        return None
    try:
        url = f"https://api.bcb.gov.br/dados/serie/bcdata.sgs.{sgs_code}/dados/ultimos/1?formato=json"
        r = requests.get(url, timeout=10)
        r.raise_for_status()
        j = r.json()
        if not j:
            return None
        val = float(str(j[0]["valor"]).replace(",", "."))
        return val / 100.0
    except Exception:
        return None


# --------------------------
# Barrier probabilities (continuous approximation + discrete correction)
# --------------------------

BETA_DISCRETE = 0.5826  # Broadie-Glasserman-Kou continuity correction

def prob_hit_down_barrier_continuous(S0: float, B: float, T: float, sigma: float, r: float = 0.0, q: float = 0.0) -> float:
    """
    P( min_{t in [0,T]} S_t <= B ) under GBM with drift r-q.
    Uses reflection principle on log-price.
    """
    if S0 <= 0 or B <= 0 or T <= 0 or sigma <= 0:
        return 0.0
    if B >= S0:
        return 1.0

    x0 = math.log(S0)
    a = math.log(B)
    mu = (r - q - 0.5 * sigma * sigma)  # drift of log
    sT = sigma * math.sqrt(T)

    z1 = (a - x0 - mu * T) / sT
    z2 = (a - x0 + mu * T) / sT

    # standard normal CDF
    N = lambda z: 0.5 * (1.0 + math.erf(z / math.sqrt(2.0)))

    # P(hit) = N(z1) + exp(2*mu*(a-x0)/sigma^2) * N(z2)
    expo = math.exp(2.0 * mu * (a - x0) / (sigma * sigma))
    p = N(z1) + expo * N(z2)
    return float(max(0.0, min(1.0, p)))

def prob_hit_down_barrier_discrete_close(S0: float, B: float, T: float, sigma: float, steps: int = 252, r: float = 0.0, q: float = 0.0) -> float:
    """
    Approximate daily-close monitored barrier by continuity correction:
      B_adj = B * exp(-beta * sigma * sqrt(dt))
    Then use continuous hit probability.
    """
    if steps <= 0:
        steps = 252
    dt_ = T / steps
    if dt_ <= 0:
        return 0.0
    B_adj = B * math.exp(-BETA_DISCRETE * sigma * math.sqrt(dt_))
    return prob_hit_down_barrier_continuous(S0, B_adj, T, sigma, r=r, q=q)

def prob_below_level_at_T(S0: float, L: float, T: float, sigma: float, r: float = 0.0, q: float = 0.0) -> float:
    """P(S_T <= L) under GBM"""
    if S0 <= 0 or L <= 0 or T <= 0 or sigma <= 0:
        return 0.0
    mu = (r - q - 0.5 * sigma * sigma)
    z = (math.log(L / S0) - mu * T) / (sigma * math.sqrt(T))
    return float(0.5 * (1.0 + math.erf(z / math.sqrt(2.0))))


# --------------------------
# SI_D_SEDE parsing
# --------------------------

def iter_sede_lines(series_path: Path):
    if series_path.suffix.lower() == ".zip":
        import zipfile
        with zipfile.ZipFile(series_path) as z:
            # assume first txt
            name = [n for n in z.namelist() if n.lower().endswith(".txt")]
            if not name:
                name = z.namelist()
            raw = z.read(name[0])
        text = raw.decode("latin-1", errors="replace")
    else:
        text = series_path.read_text(encoding="latin-1", errors="replace")
    for line in text.splitlines():
        if line.startswith("02|"):
            yield line

def build_option_universe(series_path: Path) -> List[Dict]:
    """
    Returns list of dicts:
      root4, kind('P' for put), opt_ticker, strike, expiry(yyyymmdd)
    """
    out = []
    for line in iter_sede_lines(series_path):
        parts = line.split("|")
        if len(parts) < 18:
            continue
        kind_txt = (parts[3] or "").strip().upper()
        # underlying symbol is parts[6]
        root4 = _first4_alnum(parts[6])
        opt_ticker = (parts[13] or "").strip().replace(" ", "")
        try:
            strike = float(parts[16])
        except Exception:
            continue
        try:
            expiry = int(parts[17])
        except Exception:
            continue
        kind = "P" if "VENDA" in kind_txt else ("C" if "COMPRA" in kind_txt else "")
        if not kind or not opt_ticker or not root4:
            continue
        out.append({"root4": root4, "kind": kind, "ticker": opt_ticker, "strike": strike, "expiry": expiry})
    return out


def pick_puts_for_operation(universe: List[Dict], root4: str, target_date: dt.date, barrier_price: float,
                            strike_window_pct: float = 0.10, min_strikes: int = 7) -> Tuple[int, List[Dict]]:
    """
    Choose expiry closest to target_date (preferring >= target_date), and strikes near barrier_price.
    Returns (expiry_yyyymmdd, list_of_options).
    """
    root4 = (root4 or "").upper()
    puts = [u for u in universe if u["root4"] == root4 and u["kind"] == "P"]
    if not puts:
        return (0, [])
    target = int(target_date.strftime("%Y%m%d"))

    # pick expiry: smallest positive (expiry-target), else closest abs
    expiries = sorted(set(u["expiry"] for u in puts))
    after = [e for e in expiries if e >= target]
    if after:
        expiry = min(after, key=lambda e: (e - target))
    else:
        expiry = min(expiries, key=lambda e: abs(e - target))

    candidates = [u for u in puts if u["expiry"] == expiry]
    if not candidates:
        return (expiry, [])

    # select strikes around barrier
    def select(window):
        low = barrier_price * (1 - window)
        high = barrier_price * (1 + window)
        c = [u for u in candidates if low <= u["strike"] <= high]
        c.sort(key=lambda u: u["strike"])
        return c

    c = select(strike_window_pct)
    if len(c) < min_strikes:
        c = select(max(strike_window_pct, 0.20))
    # if still few, take nearest strikes overall
    if len(c) < min_strikes:
        candidates.sort(key=lambda u: abs(u["strike"] - barrier_price))
        c = sorted(candidates[:min_strikes], key=lambda u: u["strike"])

    return expiry, c


# --------------------------
# Excel IO
# --------------------------

def read_operations(excel_path: Path, sheet_name: str = "Operações - DATA") -> List[Dict]:
    wb = load_workbook(excel_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"Aba '{sheet_name}' não encontrada.")
    ws = wb[sheet_name]
    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is None:
            continue
        headers[_norm(str(v))] = c

    # map required columns with fuzzy matches
    def col(*names):
        for n in names:
            n2 = _norm(n)
            if n2 in headers:
                return headers[n2]
        # fuzzy contains
        for k, idx in headers.items():
            for n in names:
                if _norm(n) in k:
                    return idx
        return None

    c_codigo = col("Código", "Codigo")
    c_ativo = col("Ativo")
    c_fix = col("Data de Fixing", "Fixing")
    c_venc = col("Data de vencimento", "Vencimento")
    c_ko = col("Knockout (%)", "Barreira (%)", "Barreira", "Knockout")
    c_spot = col("Spot S0", "Spot", "S0")

    if not c_ativo or not c_ko or not c_fix:
        raise RuntimeError("Colunas mínimas não encontradas (Ativo, Knockout/Barreira e Data de Fixing).")

    ops = []
    for r in range(2, ws.max_row + 1):
        ativo = ws.cell(r, c_ativo).value
        if not ativo:
            continue
        ativo = str(ativo).strip().upper()
        root4 = _first4_alnum(ativo)
        fix = _to_date(ws.cell(r, c_fix).value)
        venc = _to_date(ws.cell(r, c_venc).value) if c_venc else None
        ko = _pct_to_frac(ws.cell(r, c_ko).value)
        spot = _to_float(ws.cell(r, c_spot).value) if c_spot else None
        codigo = ws.cell(r, c_codigo).value if c_codigo else None

        ops.append({
            "row": r,
            "codigo": codigo,
            "ativo": ativo,
            "root4": root4,
            "fixing": fix,
            "venc": venc,
            "ko_frac": ko,
            "spot": spot,
        })
    return ops

def write_rtd_tickers(excel_path: Path, ops: List[Dict], picks: Dict[int, Dict]):
    wb = load_workbook(excel_path)
    ws = _safe_sheet(wb, "RTD_TICKERS")
    _clear_sheet(ws)

    ws.append(["Codigo", "Ativo", "Spot S0", "KO (%)", "Barreira (R$)", "Target Date", "Expiry (yyyymmdd)", "Tickers (1 por linha abaixo)"])

    out_rows = []
    distinct = []
    seen = set()

    for op in ops:
        r = op["row"]
        p = picks.get(r, {})
        expiry = p.get("expiry", 0)
        barrier = p.get("barrier", None)
        tdate = p.get("target_date", None)
        tickers = p.get("tickers", [])
        ws.append([op["codigo"], op["ativo"], op.get("spot"), op.get("ko_frac"), barrier,
                   tdate.strftime("%Y-%m-%d") if tdate else "", expiry, ""])
        # list tickers under it
        for t in tickers:
            ws.append(["", "", "", "", "", "", "", t])
            if t not in seen:
                distinct.append(t)
                seen.add(t)

    # also a compact list
    ws2 = _safe_sheet(wb, "ASSINAR_NO_PROFIT")
    _clear_sheet(ws2)
    ws2.append(["Cole esta coluna no Profit (1 ticker por linha)"])
    for t in distinct:
        ws2.append([t])

    wb.save(excel_path)


def read_rtd_export(excel_path: Path, sheet_name: str = "RTD_EXPORT") -> List[Dict]:
    wb = load_workbook(excel_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    if ws.max_row < 2 or ws.max_column < 2:
        return []

    headers = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is None:
            continue
        headers[_norm(str(v))] = c

    def col(*names):
        for n in names:
            n2 = _norm(n)
            if n2 in headers:
                return headers[n2]
        for k, idx in headers.items():
            for n in names:
                if _norm(n) in k:
                    return idx
        return None

    c_asset = col("Ativo", "Asset")
    if not c_asset:
        return []

    c_strike = col("Strike", "PEX")
    c_exp = col("Vencimento", "VEN")
    c_bid = col("Of. Compra", "OCP", "Bid")
    c_ask = col("Of. Venda", "OVD", "Ask")
    c_oi = col("Cont. Abertos", "CAB", "OI")
    c_vol = col("Volume", "VOL")
    c_vib = col("VI Bid", "VIB")
    c_via = col("VI Ask", "VIA")
    c_iv = col("Volt. Implícita", "Volatilidade Implícita - Opções", "Volatilidade Implícita")
    c_delta = col("Delta", "DELTA")

    rows = []
    for r in range(2, ws.max_row + 1):
        asset = ws.cell(r, c_asset).value
        if not asset:
            continue
        asset = str(asset).strip().upper().replace(" ", "")
        # ignore underlying stocks without digits? We'll keep all; later filter by needed tickers.
        strike = _to_float(ws.cell(r, c_strike).value) if c_strike else None
        exp = ws.cell(r, c_exp).value if c_exp else None
        exp_date = _to_date(exp)
        # if exp might come as yyyymmdd
        if exp_date is None:
            try:
                exp_int = int(str(exp).strip())
                exp_date = dt.datetime.strptime(str(exp_int), "%Y%m%d").date()
            except Exception:
                exp_date = None

        bid = _to_float(ws.cell(r, c_bid).value) if c_bid else None
        ask = _to_float(ws.cell(r, c_ask).value) if c_ask else None
        oi = _to_float(ws.cell(r, c_oi).value) if c_oi else None
        vol = _to_float(ws.cell(r, c_vol).value) if c_vol else None

        vib = _to_float(ws.cell(r, c_vib).value) if c_vib else None
        via = _to_float(ws.cell(r, c_via).value) if c_via else None

        iv_any = _to_float(ws.cell(r, c_iv).value) if c_iv else None
        delta = _to_float(ws.cell(r, c_delta).value) if c_delta else None

        # implied vol in decimal
        iv = None
        if vib and via:
            iv = 0.5 * (vib + via)
        elif iv_any is not None:
            iv = iv_any
        # convert % to decimal
        if iv is not None and iv > 2.0:
            iv = iv / 100.0

        rows.append({
            "asset": asset,
            "root4": _first4_alnum(asset),
            "expiry": exp_date,
            "strike": strike,
            "bid": bid,
            "ask": ask,
            "oi": oi,
            "vol": vol,
            "iv": iv,
            "delta": delta
        })
    return rows


# --------------------------
# Smile fit and probability per operation
# --------------------------

def fit_smile_sigma_at_level(option_rows: List[Dict], root4: str, expiry: dt.date, level_price: float,
                             min_oi: float = 50, min_vol: float = 10, max_spread_pct: float = 0.50,
                             delta_min: float = -0.60, delta_max: float = -0.05) -> Tuple[Optional[float], Dict]:
    """
    Fit a simple quadratic smile (iv vs log-moneyness) and return sigma(level).
    Assumes puts (negative delta). Uses liquidity filters.
    """
    root4 = root4.upper()
    rows = [r for r in option_rows if r["root4"] == root4 and r["expiry"] == expiry and r["iv"] is not None and r["strike"]]
    if not rows:
        return None, {"reason": "sem_rows_para_expiry"}

    # liquidity filters
    filt = []
    for r in rows:
        bid, ask = r.get("bid"), r.get("ask")
        if bid is not None and ask is not None and bid > 0 and ask > 0:
            spread = (ask - bid) / max(1e-9, 0.5*(ask+bid))
            if spread > max_spread_pct:
                continue
        if r.get("oi") is not None and r["oi"] < min_oi:
            continue
        if r.get("vol") is not None and r["vol"] < min_vol:
            continue
        if r.get("delta") is not None:
            if not (delta_min <= r["delta"] <= delta_max):
                continue
        filt.append(r)

    if len(filt) < 5:
        # relax if too few
        filt = rows

    strikes = np.array([r["strike"] for r in filt], dtype=float)
    ivs = np.array([r["iv"] for r in filt], dtype=float)

    # log-moneyness relative to level_price (proxy for spot region)
    x = np.log(strikes / max(1e-9, level_price))
    # weights: prefer near level and smaller spread
    w = np.exp(- (x/0.15)**2 )
    if np.all(np.isfinite(w)) and np.sum(w) > 0:
        pass
    else:
        w = None

    # fit quadratic
    try:
        if w is not None:
            coef = np.polyfit(x, ivs, deg=2, w=w)
        else:
            coef = np.polyfit(x, ivs, deg=2)
        a, b, c = coef
        sigma = float(a*0.0*0.0 + b*0.0 + c)  # at x=0 (K=level_price)
        # clip to sane range
        sigma = max(0.01, min(5.0, sigma))
        info = {"n": int(len(filt)), "coef": [float(a), float(b), float(c)]}
        return sigma, info
    except Exception as e:
        return None, {"reason": f"polyfit_falhou: {e}"}


def ensure_result_columns(excel_path: Path, sheet_name: str = "Operações - DATA") -> Dict[str, int]:
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]
    header = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is None:
            continue
        header[_norm(str(v))] = c

    def get_or_add(col_name: str) -> int:
        key = _norm(col_name)
        if key in header:
            return header[key]
        new_c = ws.max_column + 1
        ws.cell(row=1, column=new_c).value = col_name
        header[key] = new_c
        return new_c

    cols = {
        "Barreira (R$)": get_or_add("Barreira (R$)"),
        "σ_barreira (a.a.)": get_or_add("σ_barreira (a.a.)"),
        "P(KO tocar)": get_or_add("P(KO tocar)"),
        "P(abaixo no fixing)": get_or_add("P(abaixo no fixing)"),
        "r (a.a.)": get_or_add("r (a.a.)"),
        "Expiry usada": get_or_add("Expiry usada"),
        "Smile info": get_or_add("Smile info"),
    }
    wb.save(excel_path)
    return cols


def write_results(excel_path: Path, ops: List[Dict], results: Dict[int, Dict]):
    cols = ensure_result_columns(excel_path)
    wb = load_workbook(excel_path)
    ws = wb["Operações - DATA"]

    # write into operations
    for op in ops:
        r = op["row"]
        res = results.get(r)
        if not res:
            continue
        ws.cell(r, cols["Barreira (R$)"]).value = res.get("barrier")
        ws.cell(r, cols["σ_barreira (a.a.)"]).value = res.get("sigma")
        ws.cell(r, cols["P(KO tocar)"]).value = res.get("p_hit")
        ws.cell(r, cols["P(abaixo no fixing)"]).value = res.get("p_term")
        ws.cell(r, cols["r (a.a.)"]).value = res.get("r")
        ws.cell(r, cols["Expiry usada"]).value = res.get("expiry")
        ws.cell(r, cols["Smile info"]).value = json.dumps(res.get("smile_info", {}), ensure_ascii=False)

    # also a results sheet
    ws_out = _safe_sheet(wb, "RESULTADOS")
    _clear_sheet(ws_out)
    ws_out.append(["Codigo", "Ativo", "Spot", "KO(%)", "Barreira(R$)", "Fixing", "Expiry", "sigma", "P_hit_close", "P_term", "r", "smile_info"])
    for op in ops:
        r = op["row"]
        res = results.get(r)
        if not res:
            continue
        ws_out.append([op.get("codigo"), op.get("ativo"), op.get("spot"), op.get("ko_frac"), res.get("barrier"),
                       op.get("fixing").isoformat() if op.get("fixing") else "",
                       res.get("expiry"), res.get("sigma"), res.get("p_hit"), res.get("p_term"), res.get("r"),
                       json.dumps(res.get("smile_info", {}), ensure_ascii=False)])
    wb.save(excel_path)


# --------------------------
# Main
# --------------------------

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--series_file", required=True)
    parser.add_argument("--strike_window_pct", type=float, default=0.10)
    parser.add_argument("--min_oi", type=float, default=50)
    parser.add_argument("--min_vol", type=float, default=10)
    parser.add_argument("--max_spread_pct", type=float, default=0.50)
    parser.add_argument("--delta_min", type=float, default=-0.60)
    parser.add_argument("--delta_max", type=float, default=-0.05)
    parser.add_argument("--sgs_code", type=int, default=11)
    args = parser.parse_args()

    excel_path = Path(args.workbook)
    series_path = Path(args.series_file)
    if not excel_path.exists():
        raise SystemExit(f"Workbook não encontrado: {excel_path}")
    if not series_path.exists():
        raise SystemExit(f"Series file não encontrado: {series_path}")

    # 1) Read operations
    ops = read_operations(excel_path)
    if not ops:
        print("Nenhuma operação encontrada na aba 'Operações - DATA'.")
        return

    # need spot + knockout to compute barrier to pick strikes
    universe = build_option_universe(series_path)

    picks = {}
    missing_spot = 0
    for op in ops:
        if op.get("fixing") is None or op.get("ko_frac") is None:
            continue
        spot = op.get("spot")
        if not spot or spot <= 0:
            missing_spot += 1
            continue
        barrier = spot * op["ko_frac"]
        target_date = op.get("venc") or op.get("fixing")
        expiry, opt_list = pick_puts_for_operation(
            universe, op["root4"], target_date, barrier,
            strike_window_pct=args.strike_window_pct
        )
        tickers = [o["ticker"] for o in opt_list]
        picks[op["row"]] = {"expiry": expiry, "barrier": barrier, "target_date": target_date, "tickers": tickers}

    # Write RTD_TICKERS every time (so you always have the list)
    write_rtd_tickers(excel_path, ops, picks)

    if missing_spot > 0:
        print(f"⚠️  {missing_spot} operações sem Spot S0. Preencha Spot S0 (via RTD) para incluir no cálculo/seleção de opções.")

    # 2) Read RTD_EXPORT
    export_rows = read_rtd_export(excel_path)
    if not export_rows:
        print("Aba RTD_EXPORT vazia ou não encontrada.")
        print("✅  Geramos RTD_TICKERS/ASSINAR_NO_PROFIT. Cole no Profit, exporte RTD/DDE para RTD_EXPORT, SALVE a planilha e rode novamente.")
        return

    # 3) Compute rate
    r = _sgs_daily_rate(args.sgs_code) or 0.0

    # 4) For each operation, fit smile and compute probabilities
    results = {}
    for op in ops:
        if op.get("fixing") is None or op.get("ko_frac") is None:
            continue
        spot = op.get("spot")
        if not spot or spot <= 0:
            continue
        barrier = spot * op["ko_frac"]
        target_date = op.get("venc") or op.get("fixing")
        # determine expiry chosen before
        expiry_yyyymmdd = picks.get(op["row"], {}).get("expiry", 0)
        expiry_date = None
        if expiry_yyyymmdd:
            try:
                expiry_date = dt.datetime.strptime(str(expiry_yyyymmdd), "%Y%m%d").date()
            except Exception:
                expiry_date = None
        if expiry_date is None:
            # fallback: use nearest expiry in export
            exp_set = sorted({r2["expiry"] for r2 in export_rows if r2["root4"] == op["root4"] and r2["expiry"]})
            if exp_set:
                expiry_date = min(exp_set, key=lambda d: abs((d - target_date).days))
        if expiry_date is None:
            continue

        sigma, sinfo = fit_smile_sigma_at_level(
            export_rows, op["root4"], expiry_date, barrier,
            min_oi=args.min_oi, min_vol=args.min_vol, max_spread_pct=args.max_spread_pct,
            delta_min=args.delta_min, delta_max=args.delta_max
        )
        if sigma is None:
            continue

        # T from today? use Prazo (Dias) column if exists? We'll approximate using business days between now and fixing:
        # Use calendar-free: days/252, where days = max(1, (fixing - today).days)
        today = dt.date.today()
        days = max(1, (op["fixing"] - today).days) if op["fixing"] else 1
        T = days / 252.0

        p_hit = prob_hit_down_barrier_discrete_close(spot, barrier, T, sigma, steps=252, r=r, q=0.0)
        p_term = prob_below_level_at_T(spot, barrier, T, sigma, r=r, q=0.0)

        results[op["row"]] = {
            "barrier": barrier,
            "sigma": sigma,
            "p_hit": p_hit,
            "p_term": p_term,
            "r": r,
            "expiry": expiry_date.strftime("%Y-%m-%d"),
            "smile_info": sinfo
        }

    write_results(excel_path, ops, results)
    print("✅ RESULTADOS calculados e gravados (aba RESULTADOS e colunas na Operações - DATA).")
    print("Obs.: modelo considera rompimento por FECHAMENTO (P_hit) e barreira DOWN (<= barreira).")


if __name__ == "__main__":
    # IDLE-friendly: auto-detect workbook + series in same folder if no args
    if len(sys.argv) == 1:
        here = Path(__file__).resolve().parent
        # workbook preference
        wb = None
        for name in ["Modelo_Prob_Barreira_clean.xlsx", "Modelo_Prob_Barreira_v3_com_operacoes.xlsx", "Operações - DATA.xlsx"]:
            p = here / name
            if p.exists():
                wb = p
                break
        if wb is None:
            # any xlsx with Prob_Barreira in name
            cand = list(here.glob("*Prob*Barreira*.xlsx"))
            wb = cand[0] if cand else None

        sf = None
        # zip or txt
        cand = list(here.glob("SI_D_SEDE*.zip")) + list(here.glob("SI_D_SEDE*.txt"))
        sf = cand[0] if cand else None

        if wb is None or sf is None:
            print("❌ Não encontrei automaticamente workbook e/ou SI_D_SEDE na mesma pasta do .py.")
            print("Execute via terminal com argumentos:")
            print('  py calcula_prob_barreira_v7_noCOM.py --workbook "C:\\path\\to\\Model.xlsx" --series_file "C:\\path\\to\\SI_D_SEDE.zip"')
            sys.exit(1)

        sys.argv += ["--workbook", str(wb), "--series_file", str(sf)]
        print(f"Modo IDLE: workbook = {wb}")
        print(f"Modo IDLE: series_file = {sf}")

    main()
